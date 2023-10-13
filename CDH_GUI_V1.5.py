# import essential libraries 
import os
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox

###
cwd = os.getcwd()
## input typhoon name and make file
root = Tk()
root.title("CDH Downloader")
root.geometry("480x360") 
#mainframe = ttk.Frame(root, padding = "300 240 300 240")
#mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
root.columnconfigure(0, weight=0)
root.rowconfigure(0, weight=0)
###########################################
menubar = Menu(root)
root.config(menu=menubar)

# create the file_menu
file_menu = Menu(
    menubar,
    tearoff=0
)

# add menu items to the File menu
file_menu.add_separator()

# add Exit menu item
file_menu.add_command(
    label='Exit',
    command=root.destroy
)

# add the File menu to the menubar
menubar.add_cascade(
    label="File",
    menu=file_menu
)
# create the Help menu
help_menu = Menu(
    menubar,
    tearoff=0
)

#################################
def openHelpWindow():
    help_window = Toplevel(root)
    # Toplevel widget
    help_window.title("About")
    # sets the geometry of toplevel
    help_window.geometry("250x120")
    # A Label widget to show in toplevel
    Label(help_window,
          text ="\n\nĐược tạo bởi Đào Anh Công\n - Đài KTTV Bắc Trung Bộ")\
          .pack()

help_menu.add_command(label='About...',command = openHelpWindow)
#################################

# add the Help menu to the menubar
menubar.add_cascade(
    label="Help",
    menu=help_menu
)
###########################################
# pick region
ttk.Label(root, text="Chọn khu vực:").\
                 place(x=10,y =25)
regions = ['Tây Bắc','Việt Bắc','Đông Bắc','Đồng Bằng Bắc Bộ','Bắc Trung Bộ',\
           'Trung Trung Bộ','Nam Trung Bộ','Tây Nguyên','Nam Bộ']
region = StringVar()
region.set('Bắc Trung Bộ')
drop_region = OptionMenu(root,region,*regions)
drop_region.place(x=110,y =20)
###########################################
# pick observation method
ttk.Label(root, text="Chọn loại số liệu:").\
                 place(x=10,y =60)
obs_methods = ['Quan trắc khí tượng','Quan trắc thủy văn','Mưa tự động',\
           'Số liệu hồ chứa','Quan trắc biển']
method = StringVar()
method.set('Quan trắc khí tượng')
drop_method = OptionMenu(root,method,*obs_methods)
drop_method.place(x=110,y =55)
##########################################
import datetime as dt
now = str(dt.datetime.now().date())
now1 = str(dt.datetime.now().time())[0:2]
#print (now)
year1 = int(now[0:4])
month1 = int(now[5:7])
day1 = int(now[8:10])
#print (year1, month1, day1)
####
hours = []
for i in range(0,24):
   if (i<10):
      hh = '0'+str(i)
   else:
      hh = str(i)
   hours.append(hh)
start_hour = StringVar()
start_hour.set('00')
drop_start_hour = OptionMenu(root,start_hour,*hours)
drop_start_hour.place(x=60,y =125)
#begin_hour = start_hour.get()

end_hour = StringVar()
end_hour.set(now1)
drop_end_hour = OptionMenu(root,end_hour,*hours)
drop_end_hour.place(x=270,y =125)

var_s = []
#finish_hour = end_hour.get()
################### Chọn biến
class Checkbar(Frame):
   def __init__(self, parent=None, picks=[], side=LEFT, anchor=W):
      Frame.__init__(self, parent)
      self.vars = []
      for pick in picks:
         var = IntVar()
         chk = Checkbutton(self, text=pick, variable=var)
         chk.pack(side=side, anchor=anchor, expand=NO)
         self.vars.append(var)
   def state(self):
      return map((lambda var: var.get()), self.vars)
   
def var_switcher(var):
    switcher = {
        'Nhiệt độ': 'AirTemp/Manual.O',
        'Lượng mưa': 'Precip/Manual.O',
        'Độ ẩm': 'RH/Manual.O',
        'Hướng gió':'WDir/Manual.O',
        'Tốc độ gió' : 'WSpeed/Manual.O',
        'Khí áp': 'P_SL/Manual.O',
        'Lượng mây': 'Ccover',
        'Bốc hơi': 'Evap/Manual.O',
        'Tổng mưa 24h': 'Precip/Manual.24h.O',
        'Mực nước': 'S/Manual.O',
        'Lưu Lượng': 'Q/Manual.O',
        'Mưa tự động 6h': 'Precip/6h.total',
        'Mưa tự động 1h': 'Precip/h.total',
        'Mực nước hồ': 'ResLVL',
        'Lưu lượng hồ' : 'Q',
        'Tmax':'Airtemp/Manual.tx',
        'Tmin':'Airtemp/Manual.tn',
        'Độ mặn':'SALT/Manual.O',
        'Mực nước biển':'WL/Manual.O',
        'Trạng thái biển':'SeaState/Manual.O',
        'Hướng sóng':'WAVEDir/Manual.O',
        'Độ cao sóng':'WAVEHeight/Manual.O',
        'Nhiệt độ nước':'WT/Manual.O'
        }
    return switcher.get(var,"Không đúng biến")    

def var_turn(var):
    switcher = {
        'AirTemp/Manual.O':'Airtemp',
        'Precip/Manual.O':'Precip',
        'RH/Manual.O':'RH',
        'WDir/Manual.O':'WDir',
        'WSpeed/Manual.O':'Ws',
        'P_SL/Manual.O':'P_SL',
        'S/Manual.O':'S',
        'Q/Manual.O':'Q',
        'Precip/6h.total':'Precip_6h',
        'Precip/h.total':'Precip_1h',
        'ResLVL':'ResLVL',
        'Q':'Q_res',
        'Airtemp/Manual.tx':'Tmax',
        'Airtemp/Manual.tn':'Tmin',
        'SALT/Manual.O':'SALT',
        'WL/Manual.O':'WL',
        'SeaState/Manual.O':'SeaState',
        'WAVEDir/Manual.O':'WAVEDir',
        'WAVEHeight/Manual.O':'WAVEH',
        'WT/Manual.O':'WT',
        'Ccover':'Ccover',
        'Precip/Manual.24h.O':'Precip24',
        'Evap/Manual.O':'Evap'
        }
    return switcher.get(var,"Không đúng biến")    
def win_decoder(wind):
    switcher = {
        '(0) ':'0/10',
        '(1) ':'1/10',
        '(2) ':'3/10',
        '(3) ':'4/10',
        '(4) ':'5/10',
        '(5) ':'6/10',
        '(6) ':'8/10',
        '(7) ':'9/10',
        '(8) ':'10/10',
        }
    return switcher.get(wind,"Không đúng biến")
def regions_switcher(region):
    switcher = {
        'Tây Bắc' : '1',
        'Việt Bắc' : '2',
        'Đông Bắc' : '3',
        'Đồng Bằng Bắc Bộ' : '4',
        'Bắc Trung Bộ' : '5',
        'Trung Trung Bộ' : '6',
        'Nam Trung Bộ' : '7',
        'Tây Nguyên' : '8',
        'Nam Bộ' : '9'
        }
    return switcher.get(region,"Sai khu vực")
def regions_switcher_short(region):
    switcher = {
        'Tây Bắc' : 'TB',
        'Việt Bắc' : 'VB',
        'Đông Bắc' : 'DB',
        'Đồng Bằng Bắc Bộ' : 'DBBB',
        'Bắc Trung Bộ' : 'BTB',
        'Trung Trung Bộ' : 'TTB',
        'Nam Trung Bộ' : 'NTB',
        'Tây Nguyên' : 'TN',
        'Nam Bộ' : 'NB'
        }
    return switcher.get(region,"Sai khu vực")
method_choice = ''
def select_var_window():
    variables_manual_1 = ['Nhiệt độ','Lượng mưa','Độ ẩm','Hướng gió','Tốc độ gió']
    variables_manual_2 = ['Khí áp', 'Tmax','Tmin','Tổng mưa 24h','Bốc hơi','Lượng mây']
    variables_TV = ['Mực nước','Lưu Lượng']
    variables_auto = ['Mưa tự động 6h','Mưa tự động 1h']
    variables_reservoir = ['Mực nước hồ','Lưu lượng hồ']
    variables_sea_1 = ['Độ mặn','Mực nước biển','Trạng thái biển','Hướng sóng']
    variables_sea_2 = ['Độ cao sóng','Nhiệt độ nước']
    select_var_window = Toplevel(root)
    select_var_window.title("Chọn biến")
    select_var_window.geometry("480x150")
    global method_choice
    method_choice = method.get()
    print (method_choice)
    var_choice = []
    
    if (method_choice == 'Quan trắc khí tượng'):
        var_choice_1 = variables_manual_1
        var_choice_2 = variables_manual_2
    elif (method_choice == 'Quan trắc thủy văn'):
        var_choice_1 = variables_TV
        var_choice_2 = []
    elif (method_choice == 'Mưa tự động'):
        var_choice_1 = variables_auto
        var_choice_2 = []
    elif (method_choice == 'Số liệu hồ chứa'):
        var_choice_1 = variables_reservoir
        var_choice_2 = []
    elif (method_choice == 'Quan trắc biển'):
        var_choice_1 = variables_sea_1
        var_choice_2 = variables_sea_2
        
    var_choice = var_choice_1 + var_choice_2
    
    var_list_1 = Checkbar(select_var_window, var_choice_1)
    var_list_1.place(x=20, y=20)
    var_list_1.config(relief=GROOVE, bd=2)
    var_list_2 = Checkbar(select_var_window, var_choice_2)
    var_list_2.place(x=20, y=50)
    var_list_2.config(relief=GROOVE, bd=2)
    
    global begin_hour
    begin_hour = start_hour.get()
    global finish_hour
    finish_hour = end_hour.get()
    global start_date
    start_date = start_day_picker.get_date()
    global end_date
    end_date = end_day_picker.get_date()
    
    def var_states(): 
        state = list(var_list_1.state()) + list(var_list_2.state())
        global vars_name
        vars_name = []
        for i in range(len(state)):
            if (state[i] > 0):
                vars_name.append(var_choice[i])
        print (vars_name)
        global var_s
        var_s = []
        for name in vars_name:
            var_s.append(var_switcher(name))
        print (var_s)
    #for name in vars_name:
    #    var_s.append(var_switcher(name))
    Cancel_button = Button(select_var_window, text='Cancel',\
                           command=select_var_window.destroy).place(x=220, y=100)
    Ok_button = Button(select_var_window, text='Ok',\
                       command=lambda:[var_states(),select_var_window.destroy()]).\
                       place(x=180, y=100)
    return    
select_button = Button(root, text = "Chọn biến" , command = select_var_window)\
                .place(x=300,y=57)

from tkcalendar import DateEntry
ttk.Label(root, text="Bắt đầu:").\
                 place(x=10,y =130)
start_day_picker = DateEntry(root,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', \
                borderwidth=2)
start_day_picker.place(x=120,y =130)

####
ttk.Label(root, text="Kết thúc:").\
                 place(x=220,y =130)
end_day_picker = DateEntry(root,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', \
                borderwidth=2)
end_day_picker.place(x=330,y =130)

import urllib.request
import os
cwd = os.getcwd()

def find_dup(lst):
    from collections import defaultdict
    d = defaultdict(list)
    for i, elem in enumerate(lst):
        #print (i,"   ",elem)
        d[elem].append(i)
    test = []
    #print({k: v for k, v in d.items() if len(v) > 1})
    #print (d.items())
    for k, v in d.items():
        #print (k,v)
        if (len(v) > 1):
            for i in range(1,len(v)):
                #print (v)
                test.append(v[i])
    #print ("test = ",test)
    return(test)
def find_thua(lst,hour):
    test0 = []
    for i in range(len(lst)):
        if lst[i][-2:] != hour:
            test0.append(i)
    return (test0)
def get_manual_data():
    data_dir = os.path.join(cwd,"/Data/")
    output_dir = os.path.join(cwd,"/Output/")
    global region_choice
    global region_name
    region_name = region.get()
    region_choice = regions_switcher(region_name)
    region_short = regions_switcher_short(region_name)
    print (region_name,"    ",region_choice)
    global begin_hour
    begin_hour = start_hour.get()
    global finish_hour
    finish_hour = end_hour.get()
    global start_date
    start_date = start_day_picker.get_date()
    global end_date
    end_date = end_day_picker.get_date()
    
    def show_error():
        tk.messagebox.showinfo('Xảy ra lỗi','Vui lòng chọn biến')
    
    template_file_KT = os.path.join(cwd+"/Output/","Template_KT.xlsx")
    template_file_R = os.path.join(cwd+"/Output/","Template_R.xlsx")
    template_file_TV = os.path.join(cwd+"/Output/","Template_TV.xlsx")
    template_file_res = os.path.join(cwd+"/Output/","Template_Res.xlsx")
    template_fle_auto = os.path.join(cwd+"/Output/","Template_auto_R.xlsx")
    template_fle_sea = os.path.join(cwd+"/Output/","Template_Sea.xlsx")

    from shutil import copy2
    
    if (method_choice == 'Quan trắc khí tượng'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_KT_'+str(start_date)+"_"+str(end_date)+".xlsx")
        copy2(template_file_KT,output_file)
    elif (method_choice == 'Quan trắc thủy văn'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_TV_'+str(start_date)+"_"+str(end_date)+".xlsx") 
        copy2(template_file_TV,output_file)   
    elif (method_choice == 'Mưa tự động'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_R_auto_'+str(start_date)+"_"+str(end_date)+".xlsx")
        copy2(template_fle_auto,output_file)
    elif (method_choice == 'Số liệu hồ chứa'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_reservor_'+str(start_date)+"_"+str(end_date)+".xlsx")
        copy2(template_file_res,output_file)
    elif (method_choice == 'Quan trắc biển'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_SEA_'+str(start_date)+"_"+str(end_date)+".xlsx")
        copy2(template_fle_sea,output_file)
    else:
        show_error()
    print ('Đang lấy dữ liệu')
    print (var_s)
    print (start_date)
    print (end_date)
    for var in var_s:
        #var = var_switcher(name)
        if (method_choice == 'Quan trắc khí tượng' or method_choice == 'Quan trắc thủy văn'):
            if (var != 'Ccover'):
                print (var)
                url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/"+\
                 var
                print (url)
                stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
        #print (stations_file)
                urllib.request.urlretrieve(url,stations_file)
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path=5/*/"\
                +var+"&returnfields=Timestamp,Value,Observation&metadata=true&from="+\
                str(start_date)+"&to="+str(end_date)
                print (url_2)
                data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
                urllib.request.urlretrieve(url_2,data_file)
            else:
                print (var)
                url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/AirTemp/Manual.O"
                print (url)
                stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
                #print (stations_file)
                urllib.request.urlretrieve(url,stations_file)
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                   +"=gettimeseriesvalues&metadata=true&timeseriesgroup_id=4144880&"+\
                   "returnfields=Timestamp,Value,Observation&metadata=true&format=xlsx&from="+\
               str(start_date)+"&to="+str(end_date)
                print (url_2)
                data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
                urllib.request.urlretrieve(url_2,data_file)
        ###################
        elif (method_choice == 'Mưa tự động'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                 "&request=getTimeseriesList&datasource=0&format=xlsx&ts_path="+region_choice+"/*/Precip/h.total"
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                   "&request=getTimeseriesValues&datasource=0&format=xlsx&ts_path="+region_choice+"/*/"+var+\
                   "&returnfields=Timestamp,Value&metadata=true&"+\
                   "from="+str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"Auto_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        elif (method_choice == 'Quan trắc biển'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
             +"=getTimeseriesList&datasource=0&format=xlsx&ts_path="+region_choice+"/*/"+\
         var
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
               +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path="+region_choice+"/*/"\
           +var+"&returnfields=Timestamp,Value&metadata=true&from="+\
           str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"Sea_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        elif (method_choice == 'Số liệu hồ chứa'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                 "getTimeseriesList&datasource=0&format=xlsx&site_no=11*&station_name="+\
                 "Ngan%20Truoi,Huong%20Son,Ho%20Ho,Nam%20Mo,Nam%20Non,Ban%20Ve,Ban%20Ang,"+\
                 "Khe%20Bo,Nhan%20Hac%20B,Chau%20Thang,Ban%20Mong,Ho%20Song%20Muc,Hua%20Na,"+\
                 "Dong%20Van,Cua%20Dat,Xuan%20Minh,Bai%20Thuong,Cam%20Thuy%201,Ba%20Thuoc%201"+\
                 ",Ba%20Thuoc%202,Hoi%20Xuan,Thanh%20Son,Trung%20Son&ts_name=00%20-%20Original%20"+\
                 "&parametertype_name="+var
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            print ('check0')
            if (var == 'ResLVL'):
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                   "getTimeseriesValues&datasource=0&format=xlsx&ts_id=308218010,308678010,310078010,"+\
                   "309878010,310968010,311048010,311018010,309918010,311028010,310048010,311718010,"+\
                   "311698010,96483010,96248010,309908010,96533010,308088010,310058010,96313010,"+\
                   "310038010,309398010,310068010,96228010&metadata=true&"+\
                   "from="+str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            else:
                url_2="https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                       "getTimeseriesValues&datasource=0&format=xlsx&ts_id=308212010,308216010,308210010,"+\
                       "308215010,308672010,308675010,308670010,308676010,309872010,309875010,309876010,"+\
                       "310072010,310070010,310075010,309870010,310076010,310966010,310960010,310965010,"+\
                       "310962010,311010010,311015010,311012010,311040010,311046010,311042010,311016010,"+\
                       "311045010,309912010,309915010,309916010,309910010,311020010,311022010,311025010,"+\
                       "311026010,310045010,310046010,310042010,310040010,311696010,311690010,311695010,"+\
                       "311710010,311715010,311716010,311692010,311712010,96480010,96481010,96245010,"+\
                       "96246010,96251010,96478010,96243010,96486010,309900010,309905010,309902010,309906010,"+\
                       "96531010,96528010,96530010,96536010,308086010,308080010,308082010,308085010,"+\
                       "310055010,310056010,310050010,310052010,96316010,96308010,96311010,96310010,"+\
                       "310030010,310032010,310035010,310036010,309392010,309390010,309395010,309396010,"+\
                       "310062010,310066010,310065010,310060010,96226010,96231010,96223010,96225010"+\
                       "&metadata=true&"+"from="+str(start_date)+"%20"+begin_hour+":00:00&to="+\
                       str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",region_choice+var_turn(var)+"Res_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        from openpyxl import load_workbook
        sta_wb = load_workbook(stations_file, data_only=True)
        ## this is the part which read the stations name and id##
        sheet_sta = sta_wb['KiQS Result']
        last_cell = len(sheet_sta['A'])
        #print (last_cell)
        stations_name = []
        stations_id = []
        for i in range(2,last_cell+1):
            cell0 = 'A'+str(i)
            cell1 = 'B'+str(i)
            #stations_name.append(sheet_sta[cell0].value)
            #stations_id.append(sheet_sta[cell1].value)
            if (var in ['Precip/Manual.O','S/Manual.O','Q/Manual.O']):
                stations_name.append(sheet_sta[cell0].value)
                stations_id.append(sheet_sta[cell1].value)
            elif ('48' in sheet_sta[cell1].value):
                stations_name.append(sheet_sta[cell0].value)
                stations_id.append(sheet_sta[cell1].value)
        #stations_name = list(dict.fromkeys(stations_name))
        #stations_id = list(dict.fromkeys(stations_id))
        #print stations_name)
        #print (stations_id)
        data_wb = load_workbook(data_file, data_only=True)
        data = []
        obs_time = []
        sheet_names = data_wb.sheetnames
        #print (sheet_names)
        for i in range(len(stations_id)):
            stations_data = []
            obs_stamp = []
            for j in range (len(sheet_names)):
                if (stations_id[i] in sheet_names[j]):
                    sheet_name_data = sheet_names[j]
        #if (len(stations_id[i])<6):
         #   sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual.O"
        #else:
        #    sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual"
            sheet_data = data_wb[sheet_name_data]
            last_data = len(sheet_data['A'])
            #print (sheet_name_data)
            for j in range(last_data+1-14):
                if (sheet_data.cell(row = j+14, column= 1).value) is not None:
                    obs_stamp.append(str(sheet_data.cell(row = j+14, column= 1).value)[5:13])
                if (var not in ['WDir/Manual.O','WAVEDir/Manual.O','Ccover']):
                    if (sheet_data.cell(row = j+14, column= 2).value) is not None:
                        stations_data.append(sheet_data.cell(row = j+14, column= 2).value)
                else:
                    if (sheet_data.cell(row = j+14, column= 3).value) is not None:
                        stations_data.append(sheet_data.cell(row = j+14, column= 3).value)

            dup = find_dup(obs_stamp)
            #print (dup)
            if (len(dup) >= 1):
                for i in range(len(dup)-1,-1,-1):
                    #print ("i = ",i)
                    del obs_stamp[dup[i]]
                    del stations_data[dup[i]]
            #print (stations_data)
            if ('Dir' in var):
                for i in range(len(stations_data)):
                    stations_data[i] = stations_data[i].split("-")[1]
                    if (len(stations_data[i])>4):
                        stations_data[i] = 'LG'
            if (var == 'Ccover'):
                for i in range(len(stations_data)):
                    stations_data[i] = stations_data[i].split("-")[0]
                    stations_data[i] = win_decoder(stations_data[i])
            if (var == 'Airtemp/Manual.tx'):
               thua1 = find_thua(obs_stamp,'19')
               for i in range(len(thua1)-1,-1,-1):
                    del obs_stamp[i]
                    del stations_data[i]
            #if (var == 'Airtemp/Manual.tn'):
               #thua2 = find_thua(obs_stamp,'07')
               #for i in range(len(thua2)-1,-1,-1):
                    #print ("i = ",i)
                    #del obs_stamp[i]
                    #del stations_data[i]
            #print (stations_data)
            data.append(stations_data)
            obs_time.append(obs_stamp)
        #print (data)
        #print (obs_time)
        ## This is the part which write data to excel file
        data_wb = load_workbook(output_file)
        #data_wb.create_sheet(name)
        sheet_0 = data_wb[var_turn(var)]
        for i in range (len(stations_name)):
            sheet_0.cell(row= 1,column = i+2,value = stations_name[i])
            sheet_0.cell(row=2,column = i+2,value = stations_id[i].replace("-","/"))
        length = []
        for i in range (len(obs_time)):
            length.append(len(obs_time[i]))
        print(max(length))
        i_max_len = length.index(max(length))
        print (i_max_len)
    
        #print (obs_time[i_max_len])
        #print (stations_name[i_max_len])
        for j in range (len(obs_time[i_max_len])):
            sheet_0.cell(row=3+j,column = 1,value = obs_time[i_max_len][j])
            
        #for i in range (len(data)):
        #    for k in range(len(obs_time[i_max_len])):
        #        sheet_0.cell(row=k+3,column = i+2,value = -99)
        #print (data)
        for i in range (len(data)):
            for j in range(len(data[i])):
                for k in range(len(obs_time[i_max_len])):
                    if (obs_time[i_max_len][k] == obs_time[i][j]):
                        sheet_0.cell(row=k+3,column = i+2,value = data[i][j])
                    #else:
                        #sheet_0.cell(row=k+3,column = i+2,value = -99)
        data_wb.save(output_file)
    print ('Đã lấy xong dữ liệu')
    def open_file_after_done():
        MsgBox = tk.messagebox.askquestion ('Hoàn thành','Đã tải xong dữ liệu')
        if MsgBox == 'yes':
        #    root.quit()
        #else:
            os.startfile(output_file)
        
    open_file_after_done()
def open_rain_menu():
    select_rain_menu = Toplevel(root)
    select_rain_menu.title("Xem mưa tự động")
    select_rain_menu.geometry("480x480")
    now = str(dt.datetime.now().date())
    now1 = str(dt.datetime.now().time())[0:2]
    #print (now)
    year1 = int(now[0:4])
    month1 = int(now[5:7])
    day1 = int(now[8:10])
    #print (year1, month1, day1)
    ####
    hours = []
    for i in range(0,24):
        if (i<10):
            hh = '0'+str(i)
        else:
            hh = str(i)
        hours.append(hh)
    start_hour0 = StringVar()
    start_hour0.set('00')
    drop_start_hour0 = OptionMenu(select_rain_menu,start_hour0,*hours)
    drop_start_hour0.place(x=60,y=45)
    end_hour0 = StringVar()
    end_hour0.set(now1)
    drop_end_hour0 = OptionMenu(select_rain_menu,end_hour0,*hours)
    drop_end_hour0.place(x=60,y=95)
    ttk.Label(select_rain_menu, text="Bắt đầu:").\
                 place(x=10,y=50)
    start_day_picker0 = DateEntry(select_rain_menu,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', borderwidth=2)
    start_day_picker0.place(x=120,y =50)
    ####
    ttk.Label(select_rain_menu, text="Kết thúc:").\
                 place(x=10,y =100)
    end_day_picker0 = DateEntry(select_rain_menu,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', borderwidth=2)
    end_day_picker0.place(x=120,y =100)
    
    def get_rain():
        global begin_hour0
        begin_hour0 = start_hour0.get()
        global finish_hour0
        finish_hour0 = end_hour0.get()
        global start_date0
        start_date0 = start_day_picker0.get_date()
        global end_date0
        end_date0 = end_day_picker0.get_date()
        print (begin_hour0,start_date0,finish_hour0,end_date0)
        url_0= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                 "&request=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/Precip/h.total"
        print (url_0)
        stations_file =os.path.join(cwd+'/Data/',"CDH_R_station.xlsx")
        print (stations_file)
        if os.path.exists(stations_file):
            print ('No need to download station file')
        else:
            urllib.request.urlretrieve(url_0,stations_file)
        url_00= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                   "&request=getTimeseriesValues&datasource=0&format=xlsx&&timeseriesgroup_id=2998222"+\
                   "&returnfields=Timestamp,Value&metadata=true&"+\
                   "from="+str(start_date0)+"%20"+begin_hour0+":00:00&to="+str(end_date0)+"%20"+finish_hour0+":00:00"
        print (url_00)
        data_file = os.path.join(cwd+ "/Data/","CDH_R_data_"+str(begin_hour0)+str(start_date0)+"_"+str(finish_hour0)+str(end_date0)+".xlsx")
        if os.path.exists(data_file):
            print ('No need to redownload data file')
        else:
            urllib.request.urlretrieve(url_00,data_file)
        from openpyxl import load_workbook
        sta_wb = load_workbook(stations_file, data_only=True)
        ## this is the part which read the stations name and id##
        sheet_sta = sta_wb['KiQS Result']
        last_cell = len(sheet_sta['A'])
        #print (last_cell)
        global stations_name,stations_id
        stations_name = []
        stations_id = []
        for i in range(2,last_cell+1):
            cell0 = 'A'+str(i)
            cell1 = 'B'+str(i)
            stations_name.append(sheet_sta[cell0].value)
            stations_id.append(sheet_sta[cell1].value)
        data_wb = load_workbook(data_file, data_only=True)
        global data,obs_time,lats,lons
        data = []
        obs_time = []
        lats = []
        lons = []
        sheet_names = data_wb.sheetnames
        #print (sheet_names)
        for i in range(len(stations_id)):
            global stations_data,obs_stamp
            stations_data = []
            obs_stamp = []
            for j in range (len(sheet_names)):
                if (stations_id[i] in sheet_names[j]):
                    sheet_name_data = sheet_names[j]
        #if (len(stations_id[i])<6):
         #   sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual.O"
        #else:
        #    sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual"
            sheet_data = data_wb[sheet_name_data]
            last_data = len(sheet_data['A'])
            #print (sheet_name_data)
            lats.append(float(sheet_data['B3'].value))
            lons.append(float(sheet_data['B4'].value))
            for j in range(last_data+1-14):
                if (sheet_data.cell(row = j+14, column= 1).value) is not None:
                    obs_stamp.append(str(sheet_data.cell(row = j+14, column= 1).value))
                if (sheet_data.cell(row = j+14, column= 2).value) is not None:
                    stations_data.append(sheet_data.cell(row = j+14, column= 2).value)
            dup = find_dup(obs_stamp)
            #print (dup)
            if (len(dup) >= 1):
                for i in range(len(dup)-1,-1,-1):
                    #print ("i = ",i)
                    del obs_stamp[dup[i]]
                    del stations_data[dup[i]]
            data.append(stations_data)
            obs_time.append(obs_stamp)
        from shutil import copy2
        output_file = os.path.join(cwd+"/Output/",'CDH_R_auto_'+str(start_date0)+"_"+str(end_date0)+".xlsx")
        template_fle_auto = os.path.join(cwd+"/Output/","Template_auto_R.xlsx")
        copy2(template_fle_auto,output_file)
        data_wb = load_workbook(output_file)
        sheet_0 = data_wb['Precip_1h']
        sheet_0.cell(row=1,column = 1,value = 'Stations')
        sheet_0.cell(row=1,column = 2,value = 'lats')
        sheet_0.cell(row=1,column = 3,value = 'lons')
        for i in range (len(stations_name)):
            sheet_0.cell(row=i+2,column = 1,value = stations_name[i])
            sheet_0.cell(row=i+2,column = 2,value = lats[i])
            sheet_0.cell(row=i+2,column = 3,value = lons[i])
        length = []
        for i in range (len(obs_time)):
            length.append(len(obs_time[i]))
        #print(max(length))
        i_max_len = length.index(max(length))
        #print (i_max_len)
    
        #print (obs_time[i_max_len])
        #print (stations_name[i_max_len])
        for j in range (len(obs_time[i_max_len])):
            time_stamp = obs_time[i_max_len][j][5:13]
            sheet_0.cell(row=1,column = j+4,value = time_stamp)
        
        for i in range (len(data)):
            for j in range(len(data[i])):
                for k in range(len(obs_time[i_max_len])):
                    if (obs_time[i][j] == obs_time[i_max_len][k]):
                        sheet_0.cell(row=i+2,column = k+4,value = data[i][j])
        # display data                
        data_wb.save(output_file)
        data_wb.close()
        import pandas as pd
        global df_R
        df_R = pd.read_excel(output_file,sheet_name='Precip_1h',index_col=None)
        df_R = df_R[df_R['lats']>17.5]
        df_R.insert(1,'Sum',df_R.iloc[:,3:].sum(axis=1))
        df_R = df_R.sort_values(by=["Sum"],ascending=False)
        df_R = df_R.drop_duplicates(subset='Stations')
        cols = ['Stations','Sum']
        df_R['Sum'] = df_R['Sum'].apply('{:.1f}'.format)
        global output_file_2
        output_file_2 = os.path.join(cwd+"/Output/",'CDH_RR_'+str(start_date0)+"_"+str(end_date0)+".xlsx")
        df_R.to_excel(output_file_2)
        #print (cols)
        list_values = df_R.to_numpy().tolist()
        treeFrame = ttk.Frame(select_rain_menu)
        treeFrame.place(x=200,y =0)
        treeview = ttk.Treeview(treeFrame, show="headings",selectmode='browse',columns=cols, height=21,)
        for col_name in cols:
            treeview.column(col_name,anchor='w',stretch=NO, minwidth=40, width=120)
            treeview.heading(col_name, text=col_name)
        for rain_values in list_values:
            treeview.insert('', tk.END, values=rain_values) 
        treeScroll_1 = ttk.Scrollbar(treeFrame, orient="vertical", command=treeview.yview)
        treeview.configure(yscrollcommand=treeScroll_1.set)
        treeScroll_1.pack(side="right", fill="y")
        treeScroll_2 = ttk.Scrollbar(treeFrame, orient="horizontal", command=treeview.xview)
        treeview.configure(xscrollcommand=treeScroll_2.set)
        treeScroll_2.pack(side="bottom", fill="x")
        treeview.pack()
        MsgBox = tk.messagebox.askquestion ('Hoàn thành',"Đã tải xong dữ liệu mưa tự động,\n bạn có muốn mở file không ?")
        if MsgBox == 'yes':
            os.startfile(output_file_2)
        return df_R
    def View_web ():
        # This part display in map
        import folium
        from folium.features import DivIcon
        def number_DivIcon(color,number):
            global icon
            icon = DivIcon(
            icon_size=(120,36),
            icon_anchor=(12,35),
            #html='<div style="font-size: 18pt; align:center, color : black">' + '{:02d}'.format(num+1) + '</div>',
            html="""<span class="fa-stack " style="font-size: 8pt">
                    <!-- The icon that will wrap the number -->
                    <span class="fa fa-circle-o fa-stack-2x" style="color : {:s}"></span>
                    <!-- a strong element with the custom content, in this case a number -->
                    <strong class="fa-stack-1x">
                         {:.1f}  
                    </strong>
                </span>""".format(color,number))
            return icon
        #'white','lightgreen','lightblue','orange','lightred','red','purple'
        #{'lightblue', 'blue', 'lightgray', 'lightred', 'darkred', 'red', 'beige', 'orange',\
        #  'pink', 'gray', 'darkgreen', 'green', 'cadetblue', 'darkpurple', 'lightgreen', 'black', 'white', 'purple', 'darkblue'}
        def color(value0):
            if (value0 <= 0.3):
                color0 = 'lightgreen'
            elif (0.3 < value0 < 16):
                color0 = 'lightblue'
            elif (16 <= value0 < 25):
                color0 = 'blue'
            elif (25 <= value0 < 50):
                color0 = 'orange'
            elif (50 <= value0 < 100):
                color0 = 'lightred'
            elif (100 <= value0 < 200):
                color0 = 'red'
            else:
                color0 = 'purple'
            return color0
        m = folium.Map(location=[19.04,105.26],zoom_start=7.5)
        for i in range(0,len(df_R)):
            loc = [df_R.iloc[i]['lats'], df_R.iloc[i]['lons']]
            folium.Marker(
            location=loc,
            popup=df_R.iloc[i]['Stations'] +": "+df_R.iloc[i]['Sum'],
            icon=folium.Icon(color=color(float(df_R.iloc[i]['Sum'])),icon='location-pin'),
            markerColor='black',
            ).add_to(m)
            folium.Marker(
            location=loc,
            popup=df_R.iloc[i]['Stations'] +": "+df_R.iloc[i]['Sum'],
            icon= number_DivIcon('black',float(df_R.iloc[i]['Sum'])),
            ).add_to(m)
        def add_categorical_legend(folium_map, title, colors, labels):
            if len(colors) != len(labels):
                raise ValueError("colors and labels must have the same length.")

            color_by_label = dict(zip(labels, colors))
    
            legend_categories = ""     
            for label, color in color_by_label.items():
                legend_categories += f"<li><span style='background:{color}'></span>{label}</li>"
        
            legend_html = f"""
            <div id='maplegend' class='maplegend'>
            <div class='legend-title'>{title}</div>
            <div class='legend-scale'>
            <ul class='legend-labels'>
            {legend_categories}
            </ul>
            </div>
            </div>
            """
            script = f"""
            <script type="text/javascript">
            var oneTimeExecution = (function() {{
                    var executed = false;
                    return function() {{
                        if (!executed) {{
                             var checkExist = setInterval(function() {{
                                       if ((document.getElementsByClassName('leaflet-top leaflet-right').length) || (!executed)) {{
                                          document.getElementsByClassName('leaflet-top leaflet-right')[0].style.display = "flex"
                                          document.getElementsByClassName('leaflet-top leaflet-right')[0].style.flexDirection = "column"
                                          document.getElementsByClassName('leaflet-top leaflet-right')[0].innerHTML += `{legend_html}`;
                                          clearInterval(checkExist);
                                          executed = true;
                                       }}
                                    }}, 100);
                        }}
                    }};
                }})();
            oneTimeExecution()
            </script>
            """
   

            css = """

            <style type='text/css'>
            .maplegend {
            z-index:9999;
            float:right;
            background-color: rgba(255, 255, 255, 1);
            border-radius: 5px;
            border: 2px solid #bbb;
            padding: 10px;
            font-size:12px;
            positon: relative;
            }
            .maplegend .legend-title {
            text-align: left;
            margin-bottom: 5px;
            font-weight: bold;
            font-size: 90%;
            }
            .maplegend .legend-scale ul {
            margin: 0;
            margin-bottom: 5px;
            padding: 0;
            float: left;
            list-style: none;
            }
            .maplegend .legend-scale ul li {
            font-size: 80%;
            list-style: none;
            margin-left: 0;
            line-height: 18px;
            margin-bottom: 2px;
            }
            .maplegend ul.legend-labels li span {
            display: block;
            float: left;
            height: 16px;
            width: 30px;
            margin-right: 5px;
            margin-left: 0;
            border: 0px solid #ccc;
            }
            .maplegend .legend-source {
            font-size: 80%;
            color: #777;
            clear: both;
            }
            .maplegend a {
            color: #777;
            }
            </style>
            """
            folium_map.get_root().header.add_child(folium.Element(script + css))
            return folium_map
        labels0 = ['0-0.3','0.3-16 mm','15-25 mm','25-50 mm','50-100 mm','100-200 mm','> 200 mm']
        color_bar = ['lightgreen','lightblue','blue','orange','#EF8885','red','purple']
        m = add_categorical_legend(m, 'Chú giải',
                        colors = color_bar,
                        labels = labels0)
        cwd0 = os.getcwd()
        html_file = os.path.join(cwd0,"map.html")
        m.save(html_file)
        import webbrowser
        # Open website
        webbrowser.open(html_file)
        return
    def open_rain_file():
        
        return
    button_get_rain = Button(select_rain_menu, text = "Get_rain" , command = get_rain)\
         .place(x=80,y =160)
    button_web_view = Button(select_rain_menu, text = "View_web" , command = View_web)\
         .place(x=75,y =200)
    #button_web_view = Button(select_rain_menu, text = "Open_file" , command = open_rain_file)\
    #     .place(x=80,y =240)
    
    return
button = Button(root, text = "get_data" , command = get_manual_data)\
         .place(x=200,y =200)
get_rain_button = Button(root, text = "Get_auto_rain" , command = open_rain_menu)\
         .place(x=190,y =250)  

root.mainloop()
