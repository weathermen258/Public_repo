import sys,os,datetime,time
from shutil import copy2
import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
import urllib.request
import os
cwd = os.getcwd()
variables = ['AirTemp','S','Precip','Q','RH','WDir','P_SL']
#variables = ['AirTemp']

start_date = "2023-07-20"
end_date = "2023-07-21"
output_file = os.path.join(cwd+"/Output/",'Data_manual_'+start_date+"_"+end_date+".xlsx")
template_file = os.path.join(cwd+"/Output/","Template.xlsx")
def find_dup(lst):
    from collections import defaultdict
    d = defaultdict(list)
    for i, elem in enumerate(lst):
        #print (i,"   ",elem)
        d[elem].append(i)
    test = []
    #print({k: v for k, v in d.items() if len(v) > 1})
    #print (d.items())
    for k, v in d.items():
        #print (k,v)
        if (len(v) > 1):
            for i in range(1,len(v)):
                #print (v)
                test.append(v[i])
    #print ("test = ",test)
    return(test)
from shutil import copy2
copy2(template_file,output_file)
for var in variables:
    url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
             +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/"+\
         var+"/Manual.O"
    print (url)
    stations_file =os.path.join(cwd+'/Data/',var+"_station_"+start_date+"_"+end_date+".xlsx")
    #print (stations_file)
    urllib.request.urlretrieve(url,stations_file)
    url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
               +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path=5/*/"\
           +var+"/Manual.O&returnfields=Timestamp,Value&metadata=true&from="+\
           start_date+"&to="+end_date
    print (url_2)
    data_file = os.path.join(cwd+ "/Data/",var+"_data_"+start_date+"_"+end_date+".xlsx")
    urllib.request.urlretrieve(url_2,data_file)
    
    import openpyxl
    from openpyxl import Workbook
    from openpyxl import load_workbook
    sta_wb = load_workbook(stations_file, data_only=True)
    sheet_sta = sta_wb['KiQS Result']
    last_cell = len(sheet_sta['A'])
    #print (last_cell)
    stations_name = []
    stations_id = []
    for i in range(2,last_cell+1):
        cell0 = 'A'+str(i)
        stations_name.append(sheet_sta[cell0].value)
        cell1 = 'B'+str(i)
        stations_id.append(sheet_sta[cell1].value)
    #print (stations_name)
    #print (stations_id)
    data_wb = load_workbook(data_file, data_only=True)
    data = []
    obs_time = []
    sheet_names = data_wb.sheetnames
    #print (sheet_names)
    for i in range(len(stations_id)):
        stations_data = []
        obs_stamp = []
        for j in range (len(sheet_names)):
            if (stations_id[i] in sheet_names[j]):
                sheet_name_data = sheet_names[j]
        #if (len(stations_id[i])<6):
         #   sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual.O"
        #else:
        #    sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual"
        sheet_data = data_wb[sheet_name_data]
        last_data = len(sheet_data['A'])
        #print (sheet_name_data)
        for j in range(last_data+1-14):
            if (sheet_data.cell(row = j+14, column= 1).value) is not None:
                obs_stamp.append(str(sheet_data.cell(row = j+14, column= 1).value)[5:13])
            if (sheet_data.cell(row = j+14, column= 2).value) is not None:
                stations_data.append(sheet_data.cell(row = j+14, column= 2).value)
        #print (obs_stamp)
        dup = find_dup(obs_stamp)
        #print (dup)
        if (len(dup) >= 1):
            for i in range(len(dup)-1,-1,-1):
                #print ("i = ",i)
                del obs_stamp[dup[i]]
                del stations_data[dup[i]]
        data.append(stations_data)
        obs_time.append(obs_stamp)
    #print (data)
    #print (obs_time)
    data_wb = load_workbook(output_file)
    data_wb.create_sheet(var)
    sheet_0 = data_wb[var]
    for i in range (len(stations_name)):
        sheet_0.cell(row=1,column = i+2,value = stations_name[i])
        sheet_0.cell(row=2,column = i+2,value = stations_id[i])
    length = []
    for i in range (len(obs_time)):
        length.append(len(obs_time[i]))
    #print(max(length))
    i_max_len = length.index(max(length))
    #print (i_max_len)
    
    #print (obs_time[i_max_len])
    #print (stations_name[i_max_len])
    for j in range (len(obs_time[i_max_len])):
        sheet_0.cell(row=3+j,column = 1,value = obs_time[i_max_len][j])
        
    for i in range (len(data)):
        for j in range(len(data[i])):
            for k in range(len(obs_time[i_max_len])):
                if (obs_time[i][j] == obs_time[i_max_len][k]):
                    sheet_0.cell(row=k+3,column = i+2,value = data[i][j])
    data_wb.save(output_file)

print (output_file)            
        
        
        
    

