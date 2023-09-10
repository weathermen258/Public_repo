#!/lustre/cong/apps/anaconda3/bin/python
import sys,os,datetime,time
from shutil import copy2

## define the input and output excel file

obs_year = str(datetime.datetime.utcnow().year)

utc_month = datetime.datetime.utcnow().month

if (utc_month > 9):
   obs_month = str(utc_month)
else:
   obs_month = '0' + str(utc_month)

utc_day = datetime.datetime.utcnow().day

if (utc_day > 9):
   obs_day = str(utc_day)
else:
   obs_day = '0' + str(utc_day)

token_file = os.path.join('/home/cong/scripts/bc_nhiet_mua/input',obs_day+'.'+obs_month+obs_year+'.token')
print (token_file)
if os.path.isfile(token_file):
   print('token file exist')
   os._exit(1)
else:
   print ('proceed')
####
#### define path
input_file = os.path.join('/home/cong/scripts/bc_nhiet_mua/input','THANG_'+obs_month+'_'+obs_year+'_SYNOP.xlsx')
output_file = os.path.join('/home/cong/scripts/bc_nhiet_mua/input','BC_Nhiet_Mua_'+obs_day+'.'+obs_month+'.xlsx')
rain_file = os.path.join('/home/cong/scripts/bc_nhiet_mua/input','THANG_'+obs_month+'_'+obs_year+'_R.xlsx')

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook

def check_input(file_name):
   check = True
   wb = load_workbook(file_name)
   sheet0 = wb['ngay'+str(utc_day)]
   for i in range(0,22,1):
      cell_name = 'T'+str(4+i)
      if sheet0[cell_name].value is None:
         check = False
     # print (check)
   return check

#### Download the file
import urllib.request
url= 'http://10.152.40.8/Download/THANG_'+obs_month+'_'+obs_year+ '_SYNOP.xlsx'
print (url)
urllib.request.urlretrieve(url,input_file)
url_2= 'http://10.152.40.8/Download/THANG_'+obs_month+'_'+obs_year+ '_R.xlsx'
urllib.request.urlretrieve(url_2,rain_file)

### Check if input file has enough data
if check_input(input_file) == False:
   os.remove(input_file)
   os._exit(1)
else:
   print ('data is ok, continue')
   
## check if input file exist
#if os.path.isfile(input_file):
#   print ('file exists')
#   os._exit(0)
#else:
#   print ('file not exist, proceed')
   
## check if output file exist
##############################
# This part get the data
if os.path.isfile(output_file):
   print('output_File already exist')
   os.remove(output_file)
   copy2('/home/cong/scripts/bc_nhiet_mua/BC_Nhiet_Mua.xlsx',output_file)
else:
   print('output_File doesnt exist, copy new file')
   copy2('/home/cong/scripts/bc_nhiet_mua/BC_Nhiet_Mua.xlsx',output_file)
print ('checkpoint 0')
####################################
def write_data(in_file,out_file):
   wb_in = load_workbook(in_file, data_only=True)
   wb_out = load_workbook (out_file)
   
   sheet_in = wb_in['ngay'+str(utc_day)]
   sheet_out = wb_out['Auto']
   
   for i_row in range(0,22):
      for i_col in range(0,8):
         sheet_out.cell(row = 6+i_row, column= 3+i_col, value=\
                        sheet_in.cell(row = 4+i_row, column =13+i_col).value)
   print (sheet_out['M6'].value)
   wb_out.save(out_file)
   
   wb_out1 = load_workbook (out_file, data_only=True)
   sheet_out = wb_out1['Auto']
   sheet_kq = wb_out1['Ketqua']
   sheet_kq['B2'] = 'Ngày '+obs_day+' tháng '+obs_month+' năm '+obs_year
   print (sheet_out['M6'].value)
   for i_row in range(0,22):
      for i_col in range(0,3):
         sheet_kq.cell(row = 6+i_row, column= 4+i_col, value=\
                        sheet_out.cell(row = 6+i_row, column =3+i_col).value)
   
   for i_row in range(0,22):
      for i_col in range(0,5):
         sheet_kq.cell(row = 6+i_row, column= 7+i_col, value=\
                        sheet_out.cell(row = 6+i_row, column =12+i_col).value)
   wb_out1.save(out_file)
   return

#write_data(input_file,output_file)

def copy_data(in_file,out_file):
   wb_in = load_workbook(in_file, data_only=True)
   wb_out = load_workbook(out_file)

   obs1h, kq1h = [],[]
   obs7h, kq7h = [],[]
   obs13h, kq13h = [],[]
   obs19h, kq19h = [],[]
   obs_tong = []
   
   sheet_in = wb_in['ngay'+str(utc_day)]
   sheet_out = wb_out['Ketqua']
   sheet_out['B2'] = 'Ngày '+obs_day+' tháng '+obs_month+' năm '+obs_year
   for i_row in range(0,22):
      obs1h.append(sheet_in.cell(row = 4+i_row, column =16).value)
      obs7h.append(sheet_in.cell(row = 4+i_row, column =17).value)
      obs13h.append(sheet_in.cell(row = 4+i_row, column =18).value)
      obs19h.append(sheet_in.cell(row = 4+i_row, column =19).value)
      obs_tong.append(sheet_in.cell(row = 4+i_row, column =20).value)
   ###
   kq1h = [None] * 22
   kq7h = [None] * 22
   kq13h = [None] * 22
   kq19h = [None] * 22
   ###
   for i in range(len(obs1h)):
      if (obs_tong[i] != "-" and obs1h[i] != "-"):
         if (obs1h[i] == round(obs_tong[i])):
            kq1h[i] = obs_tong[i]
         else:
            kq1h[i] = obs1h[i]
      else:
         kq1h[i] = "-"
      ################# 7h
      if (obs_tong[i] != "-" and obs7h[i] != "-"):
         if (obs7h[i] == round(obs_tong[i])):
            if (obs1h[i] != "-"):
               if (obs1h[i] == obs7h[i]):
                  kq7h[i] = '-'
               else:
                  kq7h[i] = obs_tong[i] - obs1h[i]
            else:
               kq7h[i] = obs_tong[i]
         else:
            if (obs1h[i] != "-"):
               if (obs7h[i] == obs1h[i]):
                  kq7h[i] = '-'
               else:
                  kq7h[i] = obs7h[i] - obs1h[i]
            else:
               kq7h[i] = obs7h[i]
      else:
         kq7h[i] = '-'
      ################## 13h
      if (obs_tong[i] != "-" and obs13h[i] != "-"):
         if (obs13h[i] == round(obs_tong[i])):
            if (obs7h[i] != "-"):
               if (obs7h[i] == obs13h[i]):
                  kq13h[i] = '-'
               else:
                  kq13h[i] = obs_tong[i] - obs7h[i]
            else:
               kq13h[i] = obs_tong[i]
         else:
            if (obs7h[i] != "-"):
               if (obs13h[i] == obs7h[i]):
                  kq13h[i] = '-'
               else:
                  kq13h[i] = obs13h[i] - obs7h[i]
            else:
               kq13h[i] = obs13h[i]
      else:
         kq13h[i] = '-'
      ################## 19h
      if (obs_tong[i] != "-" and obs19h[i] != "-"):
         if (obs19h[i] == round(obs_tong[i])):
            if (obs13h[i] != "-"):
               if (obs13h[i] == obs19h[i]):
                  kq19h[i] = '-'
               else:
                  kq19h[i] = obs_tong[i] - obs13h[i]
            else:
               kq19h[i] = obs_tong[i]
         else:
            if (obs13h[i] != "-"):
               if (obs19h[i] == obs13h[i]):
                  kq19h[i] = '-'
               else:
                  kq19h[i] = obs19h[i] - obs13h[i]
            else:
               kq19h[i] = obs19h[i]
      else:
         kq19h[i] = '-'
      
   print(kq1h)
   print(kq7h)
   print(kq13h)
   print(kq19h)
   for i_row in range(0,22):
      for i_col in range(0,3):
         sheet_out.cell(row = 6+i_row, column= 4+i_col, value=\
                        sheet_in.cell(row = 4+i_row, column =13+i_col).value)
   for j in range (len(kq1h)):
      sheet_out.cell(row = 6+j, column= 7, value= kq1h[j])
      sheet_out.cell(row = 6+j, column= 8, value= kq7h[j])
      sheet_out.cell(row = 6+j, column= 9, value= kq13h[j])
      sheet_out.cell(row = 6+j, column= 10, value= kq19h[j])
      sheet_out.cell(row = 6+j, column= 11, value= obs_tong[j])
      
   wb_out.save(out_file)
   return
print ('check point 1')
copy_data(input_file,output_file)
####################################
# This part send mail

import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders

def send_mail(send_from,send_to,subject,text,files,username,password):
   msg = MIMEMultipart()
   msg['From'] = send_from
   recipients = send_to
   msg['To'] = ", ".join(recipients)
   msg['Date'] = formatdate(localtime = True)
   msg['Subject'] = subject
   msg.attach(MIMEText(text))
   
   for file in files:
      part = MIMEBase('application', 'octet-stream')
      part.set_payload(open(file, "rb").read())
      encoders.encode_base64(part)
      part.add_header('Content-Disposition', 'attachment', filename=file)
      msg.attach(part)
   
   #context = ssl.SSLContext(ssl.PROTOCOL_SSLv3)
   #SSL connection only working on Python 3+
   smtp = smtplib.SMTP_SSL('smtp.gmail.com', "465", context=context)
   smtp.login(username,password)
   smtp.sendmail(send_from, send_to, msg.as_string())
   smtp.quit()
#########################################

sent_mail = 'bc.nhietmua@gmail.com'
receive_mail = ['dubaokttv.dbtb@gmail.com','ducbvtvna@gmail.com','tieukttvbtb@gmail.com','tien1967@gmail.com','huankttv_btb@yahoo.com.vn','leduccuong.kttv@gmail.com','pclbnghean@gmail.com','radar.vinh.dkvbtb@gmail.com','tangandbkt@gmail.com','nguyenvanluongbtb@gmail.com','bc.nhietmua@gmail.com']
#receive_mail = 'hang3.btb@gmail.com'
subject = 'Báo cáo nhiệt mưa ngày ' + obs_day+'/'+obs_month+'/'+obs_year
body = 'Báo cáo nhiệt mưa ngày ' + obs_day+'/'+obs_month+'/'+obs_year
files = [output_file, rain_file]
username = 'bc.nhietmua@gmail.com'
password = 'olqx pxbp nvnb eduo'
####

import os
from email.message import EmailMessage
em = EmailMessage()
em['from'] = sent_mail
em['to']= receive_mail
em['subject']=subject
em.set_content(body)

context = ssl.create_default_context()
print ('teste')
send_mail(sent_mail,receive_mail,subject,body,files,username,password)
print ('pass')
## making token
copy2('/home/cong/scripts/bc_nhiet_mua/date.token',token_file)
