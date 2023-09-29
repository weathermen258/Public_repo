import os,datetime
from shutil import copy2
import urllib.request
#current working dir
cwd = os.getcwd()
variables_synop = ['AirTemp/Manual.O','Precip/Manual.O','RH/Manual.O','WDir/Manual.O',\
                   'WSpeed/Manual.O','P_SL/Manual.O','Airtemp/Manual.tx',\
                   'Airtemp/Manual.tn','Ccover', 'Precip/Manual.24h.O','Evap/Manual.O']
variables_R = ['Precip/Manual.O','Precip/Manual.24h.O']
variables_TV = ['S/Manual.O','Q/Manual.O']
variables_rain = ['Precip/6h.total','Precip/h.total']
variables_sea = ['SALT/Manual.O', 'WL/Manual.O','SeaState/Manual.O',\
                 'WAVEDir/Manual.O','WAVEHeight/Manual.O','WT/Manual.O' ]
variable_Ho = ['ResLVL','Q']
def var_switcher(var):
    switcher = {
        'AirTemp/Manual.O':'Airtemp',
        'Precip/Manual.O':'Precip',
        'RH/Manual.O':'RH',
        'WDir/Manual.O':'WDir',
        'WSpeed/Manual.O':'Ws',
        'P_SL/Manual.O':'P_SL',
        'S/Manual.O':'S',
        'Q/Manual.O':'Q',
        'Precip/6h.total':'Precip_6h',
        'Precip/h.total':'Precip_1h',
        'ResLVL':'ResLVL',
        'Q':'Q_res',
        'Airtemp/Manual.tx':'Tmax',
        'Airtemp/Manual.tn':'Tmin',
        'SALT/Manual.O':'SALT',
        'WL/Manual.O':'WL',
        'SeaState/Manual.O':'SeaState',
        'WAVEDir/Manual.O':'WAVEDir',
        'WAVEHeight/Manual.O':'WAVEH',
        'WT/Manual.O':'WT',
        'Ccover':'Ccover',
        'Precip/Manual.24h.O':'Precip24',
        'Evap/Manual.O':'Evap'
        }
    return switcher.get(var,"Không đúng biến")
###############
def win_decoder(wind):
    switcher = {
        '(0) ':'0/10',
        '(1) ':'1/10',
        '(2) ':'3/10',
        '(3) ':'4/10',
        '(4) ':'5/10',
        '(5) ':'6/10',
        '(6) ':'8/10',
        '(7) ':'9/10',
        '(8) ':'10/10',
        }
    return switcher.get(wind,"Không đúng biến")
###############
#find duplicate in cdh data
def find_dup(lst):
    from collections import defaultdict
    d = defaultdict(list)
    for i, elem in enumerate(lst):
        #print (i,"   ",elem)
        d[elem].append(i)
    test = []
    #print({k: v for k, v in d.items() if len(v) > 1})
    #print (d.items())
    for k, v in d.items():
        #print (k,v)
        if (len(v) > 1):
            for i in range(1,len(v)):
                #print (v)
                test.append(v[i])
    #print ("test = ",test)
    return(test)
def find_thua(lst,hour):
    test0 = []
    for i in range(len(lst)):
        if lst[i][-2:] != hour:
            test0.append(i)
    return (test0)
################
#get data from cdh
def get_data(variables,output_file):
    print (variables)
    for var in variables:
        if (var != 'Ccover'):
            print (var)
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
             +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/"+\
                 var
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_switcher(var)+"_station_"+start_date+"_"+end_date+".xlsx")
    #print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
               +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path=5/*/"\
               +var+"&returnfields=Timestamp,Value,Observation&metadata=true&from="+\
               start_date+"&to="+end_date
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",var_switcher(var)+"_data_"+start_date+"_"+end_date+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        else:
            print (var)
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
             +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/AirTemp/Manual.O"
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_switcher(var)+"_station_"+start_date+"_"+end_date+".xlsx")
    #print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                   +"=gettimeseriesvalues&metadata=true&timeseriesgroup_id=4144880&"+\
                   "returnfields=Timestamp,Value,Observation&metadata=true&format=xlsx&from="+\
               start_date+"&to="+end_date
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",var_switcher(var)+"_data_"+start_date+"_"+end_date+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
    
        import openpyxl
        from openpyxl import Workbook
        from openpyxl import load_workbook
        sta_wb = load_workbook(stations_file, data_only=True)
        sheet_sta = sta_wb['KiQS Result']
        last_cell = len(sheet_sta['A'])
        #print (last_cell)
        stations_name = []
        stations_id = []
        for i in range(2,last_cell+1):
            cell0 = 'A'+str(i)
            cell1 = 'B'+str(i)
            #stations_name.append(sheet_sta[cell0].value)
            #stations_id.append(sheet_sta[cell1].value)
            if (var in ['Precip/Manual.O','S/Manual.O','Q/Manual.O']):
                stations_name.append(sheet_sta[cell0].value)
                stations_id.append(sheet_sta[cell1].value)
            elif ('48' in sheet_sta[cell1].value):
                stations_name.append(sheet_sta[cell0].value)
                stations_id.append(sheet_sta[cell1].value)
            
        #print (stations_name)
        #print (stations_id)
        data_wb = load_workbook(data_file, data_only=True)
        data = []
        obs_time = []
        sheet_names = data_wb.sheetnames
        #print (sheet_names)
        for i in range(len(stations_id)):
            stations_data = []
            obs_stamp = []
            for j in range (len(sheet_names)):
                if (stations_id[i] in sheet_names[j]):
                    sheet_name_data = sheet_names[j]
            #if (len(stations_id[i])<6):
             #   sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual.O"
            #else:
            #    sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual"
            sheet_data = data_wb[sheet_name_data]
            last_data = len(sheet_data['A'])
            #print (sheet_name_data)
            for j in range(last_data+1-14):
                if (sheet_data.cell(row = j+14, column= 1).value) is not None:
                    obs_stamp.append(str(sheet_data.cell(row = j+14, column= 1).value)[5:13])
                if (var not in ['WDir/Manual.O','WAVEDir/Manual.O','Ccover']):
                    if (sheet_data.cell(row = j+14, column= 2).value) is not None:
                        stations_data.append(sheet_data.cell(row = j+14, column= 2).value)
                else:
                    if (sheet_data.cell(row = j+14, column= 3).value) is not None:
                        stations_data.append(sheet_data.cell(row = j+14, column= 3).value)
            #print (obs_stamp)
            dup = find_dup(obs_stamp)
            #print (dup)
            
            if (len(dup) >= 1):
                for i in range(len(dup)-1,-1,-1):
                    #print ("i = ",i)
                    del obs_stamp[dup[i]]
                    del stations_data[dup[i]]
            #print (stations_data)
            if ('Dir' in var):
                for i in range(len(stations_data)):
                    stations_data[i] = stations_data[i].split("-")[1]
                    if (len(stations_data[i])>4):
                        stations_data[i] = 'LG'
            if (var == 'Ccover'):
                for i in range(len(stations_data)):
                    stations_data[i] = stations_data[i].split("-")[0]
                    stations_data[i] = win_decoder(stations_data[i])
            if (var == 'Airtemp/Manual.tx'):
               thua1 = find_thua(obs_stamp,'19')
               for i in range(len(thua1)-1,-1,-1):
                    del obs_stamp[i]
                    del stations_data[i]
            #if (var == 'Airtemp/Manual.tn'):
               #thua2 = find_thua(obs_stamp,'07')
               #for i in range(len(thua2)-1,-1,-1):
                    #print ("i = ",i)
                    #del obs_stamp[i]
                    #del stations_data[i]
            #print (stations_data)
            data.append(stations_data)
            obs_time.append(obs_stamp)
        #print (data)
        #print (obs_time)
        data_wb = load_workbook(output_file)
        #data_wb.create_sheet(var_switcher(var))
        sheet_0 = data_wb[var_switcher(var)]
        for i in range (len(stations_name)):
            sheet_0.cell(row=1,column = i+2,value = stations_name[i])
            sheet_0.cell(row=2,column = i+2,value = stations_id[i].replace("-","/"))
        length = []
        for i in range (len(obs_time)):
            length.append(len(obs_time[i]))
        print(length)
        i_max_len = length.index(max(length))
        print (i_max_len)
    
        #print (obs_time[i_max_len])
        #print (stations_name[i_max_len])
        for j in range (len(obs_time[i_max_len])):
            sheet_0.cell(row=3+j,column = 1,value = obs_time[i_max_len][j])
            
        #for i in range (len(data)):
        #    for k in range(len(obs_time[i_max_len])):
        #        sheet_0.cell(row=k+3,column = i+2,value = -99)
        #print (data)
        for i in range (len(data)):
            for j in range(len(data[i])):
                for k in range(len(obs_time[i_max_len])):
                    if (obs_time[i_max_len][k] == obs_time[i][j]):
                        sheet_0.cell(row=k+3,column = i+2,value = data[i][j])
                    #else:
                        #sheet_0.cell(row=k+3,column = i+2,value = -99)
        data_wb.save(output_file)
######
####################
#define start and end date
import datetime
from datetime import timedelta
now = datetime.datetime.now().date()
end_date = str(now)
year = int(now.strftime('%Y'))
month = int(now.strftime('%m'))
start_date_obj = datetime.datetime(year,month,1) - timedelta(days=1)
start_date = start_date_obj.strftime('%Y-%m-%d')
start_date = '2023-05-30'
end_date = '2023-06-30'
print (start_date, end_date)
##################
#define output file
template_file_KT = os.path.join(cwd+"/Output/","Template_KT.xlsx")
output_file_KT = os.path.join(cwd+"/Output/",'CDH_KT_Thang_'+end_date.split("-")[1]+"_"+end_date.split("-")[0]+".xlsx")

template_file_R = os.path.join(cwd+"/Output/","Template_R.xlsx")
output_file_R = os.path.join(cwd+"/Output/",'CDH_R_Thang_'+end_date.split("-")[1]+"_"+end_date.split("-")[0]+".xlsx")

template_file_TV = os.path.join(cwd+"/Output/","Template_TV.xlsx")
output_file_TV = os.path.join(cwd+"/Output/",'CDH_TV_Thang_'+end_date.split("-")[1]+"_"+end_date.split("-")[0]+".xlsx")
from shutil import copy2
copy2(template_file_KT,output_file_KT)
copy2(template_file_R,output_file_R)
copy2(template_file_TV,output_file_TV)
#print (output_file_KT)
#get_data(variables_synop,output_file_KT)
#get_data(variables_TV,output_file_TV)
get_data(variables_R,output_file_R)

        
        
    

