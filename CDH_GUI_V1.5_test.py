# import essential libraries 
import os
from tkinter import *
import tkinter as tk
from tkinter import ttk

###
cwd = os.getcwd()
## input typhoon name and make file
root = Tk()
root.title("CDH Downloader")
root.geometry("480x360") 
#mainframe = ttk.Frame(root, padding = "300 240 300 240")
#mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
root.columnconfigure(0, weight=0)
root.rowconfigure(0, weight=0)
###########################################
menubar = Menu(root)
root.config(menu=menubar)

# create the file_menu
file_menu = Menu(
    menubar,
    tearoff=0
)

# add menu items to the File menu
file_menu.add_command(label='Thêm danh sách trạm')
file_menu.add_command(label='Open...')
file_menu.add_separator()

# add Exit menu item
file_menu.add_command(
    label='Exit',
    command=root.destroy
)

# add the File menu to the menubar
menubar.add_cascade(
    label="File",
    menu=file_menu
)
# create the Help menu
help_menu = Menu(
    menubar,
    tearoff=0
)

#################################
def openHelpWindow():
    help_window = Toplevel(root)
    # Toplevel widget
    help_window.title("About")
    # sets the geometry of toplevel
    help_window.geometry("250x120")
    # A Label widget to show in toplevel
    Label(help_window,
          text ="\n\nĐược tạo bởi Đào Anh Công\n - Đài KTTV Bắc Trung Bộ")\
          .pack()

help_menu.add_command(label='About...',command = openHelpWindow)
#################################

# add the Help menu to the menubar
menubar.add_cascade(
    label="Help",
    menu=help_menu
)
###########################################
# pick region
ttk.Label(root, text="Chọn khu vực:").\
                 place(x=10,y =25)
regions = ['Tây Bắc','Việt Bắc','Đông Bắc','Đồng Bằng Bắc Bộ','Bắc Trung Bộ',\
           'Trung Trung Bộ','Nam Trung Bộ','Tây Nguyên','Nam Bộ']
region = StringVar()
region.set('Bắc Trung Bộ')
drop_region = OptionMenu(root,region,*regions)
drop_region.place(x=110,y =20)
###########################################
# pick observation method
ttk.Label(root, text="Chọn loại số liệu:").\
                 place(x=10,y =60)
obs_methods = ['Quan trắc khí tượng','Quan trắc thủy văn','Mưa tự động',\
           'Số liệu hồ chứa','Quan trắc biển']
method = StringVar()
method.set('Quan trắc khí tượng')
drop_method = OptionMenu(root,method,*obs_methods)
drop_method.place(x=110,y =55)
##########################################
import datetime as dt
now = str(dt.datetime.now().date())
now1 = str(dt.datetime.now().time())[0:2]
#print (now)
year1 = int(now[0:4])
month1 = int(now[5:7])
day1 = int(now[8:10])
#print (year1, month1, day1)
####
hours = []
for i in range(0,24):
   if (i<10):
      hh = '0'+str(i)
   else:
      hh = str(i)
   hours.append(hh)
start_hour = StringVar()
start_hour.set('00')
drop_start_hour = OptionMenu(root,start_hour,*hours)
drop_start_hour.place(x=60,y =125)
#begin_hour = start_hour.get()

end_hour = StringVar()
end_hour.set(now1)
drop_end_hour = OptionMenu(root,end_hour,*hours)
drop_end_hour.place(x=270,y =125)

var_s = []
#finish_hour = end_hour.get()
################### Chọn biến
class Checkbar(Frame):
   def __init__(self, parent=None, picks=[], side=LEFT, anchor=W):
      Frame.__init__(self, parent)
      self.vars = []
      for pick in picks:
         var = IntVar()
         chk = Checkbutton(self, text=pick, variable=var)
         chk.pack(side=side, anchor=anchor, expand=NO)
         self.vars.append(var)
   def state(self):
      return map((lambda var: var.get()), self.vars)
   
def var_switcher(var):
    switcher = {
        'Nhiệt độ': 'AirTemp/Manual.O',
        'Lượng mưa': 'Precip/Manual.O',
        'Độ ẩm': 'RH/Manual.O',
        'Hướng gió':'WDir/Manual.O',
        'Tốc độ gió' : 'WSpeed/Manual.O',
        'Khí áp': 'P_SL/Manual.O',
        'Lượng mây': 'Ccover',
        'Bốc hơi': 'Evap/Manual.O',
        'Tổng mưa 24h': 'Precip/Manual.24h.O',
        'Mực nước': 'S/Manual.O',
        'Lưu Lượng': 'Q/Manual.O',
        'Mưa tự động 6h': 'Precip/6h.total',
        'Mưa tự động 1h': 'Precip/h.total',
        'Mực nước hồ': 'ResLVL',
        'Lưu lượng hồ' : 'Q',
        'Tmax':'Airtemp/Manual.tx',
        'Tmin':'Airtemp/Manual.tn',
        'Độ mặn':'SALT/Manual.O',
        'Mực nước biển':'WL/Manual.O',
        'Trạng thái biển':'SeaState/Manual.O',
        'Hướng sóng':'WAVEDir/Manual.O',
        'Độ cao sóng':'WAVEHeight/Manual.O',
        'Nhiệt độ nước':'WT/Manual.O'
        }
    return switcher.get(var,"Không đúng biến")    

def var_turn(var):
    switcher = {
        'AirTemp/Manual.O':'Airtemp',
        'Precip/Manual.O':'Precip',
        'RH/Manual.O':'RH',
        'WDir/Manual.O':'WDir',
        'WSpeed/Manual.O':'Ws',
        'P_SL/Manual.O':'P_SL',
        'S/Manual.O':'S',
        'Q/Manual.O':'Q',
        'Precip/6h.total':'Precip_6h',
        'Precip/h.total':'Precip_1h',
        'ResLVL':'ResLVL',
        'Q':'Q_res',
        'Airtemp/Manual.tx':'Tmax',
        'Airtemp/Manual.tn':'Tmin',
        'SALT/Manual.O':'SALT',
        'WL/Manual.O':'WL',
        'SeaState/Manual.O':'SeaState',
        'WAVEDir/Manual.O':'WAVEDir',
        'WAVEHeight/Manual.O':'WAVEH',
        'WT/Manual.O':'WT',
        'Ccover':'Ccover',
        'Precip/Manual.24h.O':'Precip24',
        'Evap/Manual.O':'Evap'
        }
    return switcher.get(var,"Không đúng biến")    
def win_decoder(wind):
    switcher = {
        '(0) ':'0/10',
        '(1) ':'1/10',
        '(2) ':'3/10',
        '(3) ':'4/10',
        '(4) ':'5/10',
        '(5) ':'6/10',
        '(6) ':'8/10',
        '(7) ':'9/10',
        '(8) ':'10/10',
        }
    return switcher.get(wind,"Không đúng biến")
def regions_switcher(region):
    switcher = {
        'Tây Bắc' : '1',
        'Việt Bắc' : '2',
        'Đông Bắc' : '3',
        'Đồng Bằng Bắc Bộ' : '4',
        'Bắc Trung Bộ' : '5',
        'Trung Trung Bộ' : '6',
        'Nam Trung Bộ' : '7',
        'Tây Nguyên' : '8',
        'Nam Bộ' : '9'
        }
    return switcher.get(region,"Sai khu vực")
def regions_switcher_short(region):
    switcher = {
        'Tây Bắc' : 'Tay_Bac',
        'Việt Bắc' : 'Viet_Bac',
        'Đông Bắc' : 'Dong_Bac',
        'Đồng Bằng Bắc Bộ' : 'Dong_Bang_Bac_Bo',
        'Bắc Trung Bộ' : 'Bac_Trung_Bo',
        'Trung Trung Bộ' : 'Trung_Trung_Bo',
        'Nam Trung Bộ' : 'Nam_Trung_Bo',
        'Tây Nguyên' : 'Tay_Nguyen',
        'Nam Bộ' : 'Nam_Bo'
        }
    return switcher.get(region,"Sai khu vực")
method_choice = ''
def select_var_window():
    variables_manual_1 = ['Nhiệt độ','Lượng mưa','Độ ẩm','Hướng gió','Tốc độ gió']
    variables_manual_2 = ['Khí áp', 'Tmax','Tmin','Tổng mưa 24h','Bốc hơi','Lượng mây']
    variables_TV = ['Mực nước','Lưu Lượng']
    variables_auto = ['Mưa tự động 6h','Mưa tự động 1h']
    variables_reservoir = ['Mực nước hồ','Lưu lượng hồ']
    variables_sea_1 = ['Độ mặn','Mực nước biển','Trạng thái biển','Hướng sóng']
    variables_sea_2 = ['Độ cao sóng','Nhiệt độ nước']
    select_var_window = Toplevel(root)
    select_var_window.title("Chọn biến")
    select_var_window.geometry("480x150")
    global method_choice
    method_choice = method.get()
    print (method_choice)
    var_choice = []
    
    if (method_choice == 'Quan trắc khí tượng'):
        var_choice_1 = variables_manual_1
        var_choice_2 = variables_manual_2
    elif (method_choice == 'Quan trắc thủy văn'):
        var_choice_1 = variables_TV
        var_choice_2 = []
    elif (method_choice == 'Mưa tự động'):
        var_choice_1 = variables_auto
        var_choice_2 = []
    elif (method_choice == 'Số liệu hồ chứa'):
        var_choice_1 = variables_reservoir
        var_choice_2 = []
    elif (method_choice == 'Quan trắc biển'):
        var_choice_1 = variables_sea_1
        var_choice_2 = variables_sea_2
        
    var_choice = var_choice_1 + var_choice_2
    
    var_list_1 = Checkbar(select_var_window, var_choice_1)
    var_list_1.place(x=20, y=20)
    var_list_1.config(relief=GROOVE, bd=2)
    var_list_2 = Checkbar(select_var_window, var_choice_2)
    var_list_2.place(x=20, y=50)
    var_list_2.config(relief=GROOVE, bd=2)
    
    global begin_hour
    begin_hour = start_hour.get()
    global finish_hour
    finish_hour = end_hour.get()
    global start_date
    start_date = start_day_picker.get_date()
    global end_date
    end_date = end_day_picker.get_date()
    
    def var_states(): 
        state = list(var_list_1.state()) + list(var_list_2.state())
        global vars_name
        vars_name = []
        for i in range(len(state)):
            if (state[i] > 0):
                vars_name.append(var_choice[i])
        print (vars_name)
        global var_s
        var_s = []
        for name in vars_name:
            var_s.append(var_switcher(name))
        print (var_s)
    #for name in vars_name:
    #    var_s.append(var_switcher(name))
    Cancel_button = Button(select_var_window, text='Cancel',\
                           command=select_var_window.destroy).place(x=220, y=100)
    Ok_button = Button(select_var_window, text='Ok',\
                       command=lambda:[var_states(),select_var_window.destroy()]).\
                       place(x=180, y=100)
    
select_button = Button(root, text = "Chọn biến" , command = select_var_window)\
                .place(x=300,y=57)

from tkcalendar import DateEntry
ttk.Label(root, text="Bắt đầu:").\
                 place(x=10,y =130)
start_day_picker = DateEntry(root,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', \
                borderwidth=2)
start_day_picker.place(x=120,y =130)

####
ttk.Label(root, text="Kết thúc:").\
                 place(x=220,y =130)
end_day_picker = DateEntry(root,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', \
                borderwidth=2)
end_day_picker.place(x=330,y =130)

import urllib.request
import os
cwd = os.getcwd()

def find_dup(lst):
    from collections import defaultdict
    d = defaultdict(list)
    for i, elem in enumerate(lst):
        #print (i,"   ",elem)
        d[elem].append(i)
    test = []
    #print({k: v for k, v in d.items() if len(v) > 1})
    #print (d.items())
    for k, v in d.items():
        #print (k,v)
        if (len(v) > 1):
            for i in range(1,len(v)):
                #print (v)
                test.append(v[i])
    #print ("test = ",test)
    return(test)
def find_thua(lst,hour):
    test0 = []
    for i in range(len(lst)):
        if lst[i][-2:] != hour:
            test0.append(i)
    return (test0)
def get_manual_data():
    data_dir = os.path.join(cwd,"/Data/")
    output_dir = os.path.join(cwd,"/Output/")
    global region_choice
    global region_name
    region_name = region.get()
    region_choice = regions_switcher(region_name)
    region_short = regions_switcher_short(region_name)
    print (region_name,"    ",region_choice)
    global begin_hour
    begin_hour = start_hour.get()
    global finish_hour
    finish_hour = end_hour.get()
    global start_date
    start_date = start_day_picker.get_date()
    global end_date
    end_date = end_day_picker.get_date()
    
    def show_error():
        tk.messagebox.showinfo('Xảy ra lỗi','Vui lòng chọn biến')
    
    template_file_KT = os.path.join(cwd+"/Output/","Template_KT.xlsx")
    template_file_R = os.path.join(cwd+"/Output/","Template_R.xlsx")
    template_file_TV = os.path.join(cwd+"/Output/","Template_TV.xlsx")

    from shutil import copy2
    
    if (method_choice == 'Quan trắc khí tượng'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_KT_'+str(start_date)+"_"+str(end_date)+".xlsx")
        copy2(template_file_KT,output_file)
    elif (method_choice == 'Quan trắc thủy văn'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_TV_'+str(start_date)+"_"+str(end_date)+".xlsx") 
        copy2(template_file_TV,output_file)   
    elif (method_choice == 'Mưa tự động'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_R_auto_'+str(start_date)+"_"+str(end_date)+".xlsx")
    elif (method_choice == 'Số liệu hồ chứa'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_reservor_'+str(start_date)+"_"+str(end_date)+".xlsx")
    elif (method_choice == 'Quan trắc biển'):
        output_file = os.path.join(cwd+"/Output/",region_short+'_CDH_SEA_'+str(start_date)+"_"+str(end_date)+".xlsx")
    else:
        show_error()
    print ('Đang lấy dữ liệu')
    print (var_s)
    print (start_date)
    print (end_date)
    for var in var_s:
        #var = var_switcher(name)
        if (method_choice == 'Quan trắc khí tượng' or method_choice == 'Quan trắc thủy văn'):
            if (var != 'Ccover'):
                print (var)
                url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/"+\
                 var
                print (url)
                stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
        #print (stations_file)
                urllib.request.urlretrieve(url,stations_file)
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path=5/*/"\
                +var+"&returnfields=Timestamp,Value,Observation&metadata=true&from="+\
                str(start_date)+"&to="+str(end_date)
                print (url_2)
                data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
                urllib.request.urlretrieve(url_2,data_file)
            else:
                print (var)
                url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/AirTemp/Manual.O"
                print (url)
                stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
                #print (stations_file)
                urllib.request.urlretrieve(url,stations_file)
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
                   +"=gettimeseriesvalues&metadata=true&timeseriesgroup_id=4144880&"+\
                   "returnfields=Timestamp,Value,Observation&metadata=true&format=xlsx&from="+\
               str(start_date)+"&to="+str(end_date)
                print (url_2)
                data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
                urllib.request.urlretrieve(url_2,data_file)
        ###################
        elif (method_choice == 'Mưa tự động'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                 "&request=getTimeseriesList&datasource=0&format=xlsx&ts_path="+region_choice+"/*/Precip/6h.total"
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                   "&request=getTimeseriesValues&datasource=0&format=xlsx&ts_path="+region_choice+"/*/"+var+\
                   "&returnfields=Timestamp,Value&metadata=true&"+\
                   "from="+str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"Auto_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        elif (method_choice == 'Quan trắc biển'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
             +"=getTimeseriesList&datasource=0&format=xlsx&ts_path="+region_choice+"/*/"+\
         var
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
               +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path="+region_choice+"/*/"\
           +var+"&returnfields=Timestamp,Value&metadata=true&from="+\
           str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",var_turn(var)+"Sea_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        elif (method_choice == 'Số liệu hồ chứa'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                 "getTimeseriesList&datasource=0&format=xlsx&site_no=11*&station_name="+\
                 "Ngan%20Truoi,Huong%20Son,Ho%20Ho,Nam%20Mo,Nam%20Non,Ban%20Ve,Ban%20Ang,"+\
                 "Khe%20Bo,Nhan%20Hac%20B,Chau%20Thang,Ban%20Mong,Ho%20Song%20Muc,Hua%20Na,"+\
                 "Dong%20Van,Cua%20Dat,Xuan%20Minh,Bai%20Thuong,Cam%20Thuy%201,Ba%20Thuoc%201"+\
                 ",Ba%20Thuoc%202,Hoi%20Xuan,Thanh%20Son,Trung%20Son&ts_name=00%20-%20Original%20"+\
                 "&parametertype_name="+var
            print (url)
            stations_file =os.path.join(cwd+'/Data/',var_turn(var)+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            print ('check0')
            if (var == 'ResLVL'):
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                   "getTimeseriesValues&datasource=0&format=xlsx&ts_id=308218010,308678010,310078010,"+\
                   "309878010,310968010,311048010,311018010,309918010,311028010,310048010,311718010,"+\
                   "311698010,96483010,96248010,309908010,96533010,308088010,310058010,96313010,"+\
                   "310038010,309398010,310068010,96228010&metadata=true&"+\
                   "from="+str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            else:
                url_2="https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                       "getTimeseriesValues&datasource=0&format=xlsx&ts_id=308212010,308216010,308210010,"+\
                       "308215010,308672010,308675010,308670010,308676010,309872010,309875010,309876010,"+\
                       "310072010,310070010,310075010,309870010,310076010,310966010,310960010,310965010,"+\
                       "310962010,311010010,311015010,311012010,311040010,311046010,311042010,311016010,"+\
                       "311045010,309912010,309915010,309916010,309910010,311020010,311022010,311025010,"+\
                       "311026010,310045010,310046010,310042010,310040010,311696010,311690010,311695010,"+\
                       "311710010,311715010,311716010,311692010,311712010,96480010,96481010,96245010,"+\
                       "96246010,96251010,96478010,96243010,96486010,309900010,309905010,309902010,309906010,"+\
                       "96531010,96528010,96530010,96536010,308086010,308080010,308082010,308085010,"+\
                       "310055010,310056010,310050010,310052010,96316010,96308010,96311010,96310010,"+\
                       "310030010,310032010,310035010,310036010,309392010,309390010,309395010,309396010,"+\
                       "310062010,310066010,310065010,310060010,96226010,96231010,96223010,96225010"+\
                       "&metadata=true&"+"from="+str(start_date)+"%20"+begin_hour+":00:00&to="+\
                       str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",region_choice+var_turn(var)+"Res_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
            
        
        
        import openpyxl
        from openpyxl import Workbook
        from openpyxl import load_workbook
        sta_wb = load_workbook(stations_file, data_only=True)
        ## this is the part which read the stations name and id##
        sheet_sta = sta_wb['KiQS Result']
        last_cell = len(sheet_sta['A'])
        #print (last_cell)
        stations_name = []
        stations_id = []
        for i in range(2,last_cell+1):
            cell0 = 'A'+str(i)
            cell1 = 'B'+str(i)
            #stations_name.append(sheet_sta[cell0].value)
            #stations_id.append(sheet_sta[cell1].value)
            if (var in ['Precip/Manual.O','S/Manual.O','Q/Manual.O']):
                stations_name.append(sheet_sta[cell0].value)
                stations_id.append(sheet_sta[cell1].value)
            elif ('48' in sheet_sta[cell1].value):
                stations_name.append(sheet_sta[cell0].value)
                stations_id.append(sheet_sta[cell1].value)
        #stations_name = list(dict.fromkeys(stations_name))
        #stations_id = list(dict.fromkeys(stations_id))
        #print stations_name)
        #print (stations_id)
        data_wb = load_workbook(data_file, data_only=True)
        data = []
        obs_time = []
        sheet_names = data_wb.sheetnames
        #print (sheet_names)
        for i in range(len(stations_id)):
            stations_data = []
            obs_stamp = []
            for j in range (len(sheet_names)):
                if (stations_id[i] in sheet_names[j]):
                    sheet_name_data = sheet_names[j]
        #if (len(stations_id[i])<6):
         #   sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual.O"
        #else:
        #    sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual"
            sheet_data = data_wb[sheet_name_data]
            last_data = len(sheet_data['A'])
            #print (sheet_name_data)
            for j in range(last_data+1-14):
                if (sheet_data.cell(row = j+14, column= 1).value) is not None:
                    obs_stamp.append(str(sheet_data.cell(row = j+14, column= 1).value)[5:13])
                if (var not in ['WDir/Manual.O','WAVEDir/Manual.O','Ccover']):
                    if (sheet_data.cell(row = j+14, column= 2).value) is not None:
                        stations_data.append(sheet_data.cell(row = j+14, column= 2).value)
                else:
                    if (sheet_data.cell(row = j+14, column= 3).value) is not None:
                        stations_data.append(sheet_data.cell(row = j+14, column= 3).value)

            dup = find_dup(obs_stamp)
            #print (dup)
            if (len(dup) >= 1):
                for i in range(len(dup)-1,-1,-1):
                    #print ("i = ",i)
                    del obs_stamp[dup[i]]
                    del stations_data[dup[i]]
            #print (stations_data)
            if ('Dir' in var):
                for i in range(len(stations_data)):
                    stations_data[i] = stations_data[i].split("-")[1]
                    if (len(stations_data[i])>4):
                        stations_data[i] = 'LG'
            if (var == 'Ccover'):
                for i in range(len(stations_data)):
                    stations_data[i] = stations_data[i].split("-")[0]
                    stations_data[i] = win_decoder(stations_data[i])
            if (var == 'Airtemp/Manual.tx'):
               thua1 = find_thua(obs_stamp,'19')
               for i in range(len(thua1)-1,-1,-1):
                    del obs_stamp[i]
                    del stations_data[i]
            #if (var == 'Airtemp/Manual.tn'):
               #thua2 = find_thua(obs_stamp,'07')
               #for i in range(len(thua2)-1,-1,-1):
                    #print ("i = ",i)
                    #del obs_stamp[i]
                    #del stations_data[i]
            #print (stations_data)
            data.append(stations_data)
            obs_time.append(obs_stamp)
        #print (data)
        #print (obs_time)
        ## This is the part which write data to excel file
        data_wb = load_workbook(output_file)
        #data_wb.create_sheet(name)
        sheet_0 = data_wb[var_turn(var)]
        for i in range (len(stations_name)):
            sheet_0.cell(row= 1,column = i+2,value = stations_name[i])
            sheet_0.cell(row=2,column = i+2,value = stations_id[i].replace("-","/"))
        length = []
        for i in range (len(obs_time)):
            length.append(len(obs_time[i]))
        print(max(length))
        i_max_len = length.index(max(length))
        print (i_max_len)
    
        #print (obs_time[i_max_len])
        #print (stations_name[i_max_len])
        for j in range (len(obs_time[i_max_len])):
            sheet_0.cell(row=3+j,column = 1,value = obs_time[i_max_len][j])
            
        #for i in range (len(data)):
        #    for k in range(len(obs_time[i_max_len])):
        #        sheet_0.cell(row=k+3,column = i+2,value = -99)
        #print (data)
        for i in range (len(data)):
            for j in range(len(data[i])):
                for k in range(len(obs_time[i_max_len])):
                    if (obs_time[i_max_len][k] == obs_time[i][j]):
                        sheet_0.cell(row=k+3,column = i+2,value = data[i][j])
                    #else:
                        #sheet_0.cell(row=k+3,column = i+2,value = -99)
        data_wb.save(output_file)
    print ('Đã lấy xong dữ liệu')
    def open_file_after_done():
        MsgBox = tk.messagebox.askquestion ('Hoàn thành','Đã tải xong dữ liệu')
        if MsgBox == 'yes':
        #    root.quit()
        #else:
            os.startfile(output_file)
        
    open_file_after_done()

button = Button(root, text = "get_data" , command = get_manual_data)\
         .place(x=200,y =200)
        

root.mainloop()
