# import essential libraries 
import os
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
import tkinter.messagebox
import babel.numbers

###
cwd = os.getcwd()
## input typhoon name and make file
root = Tk()
root.title("CDH Downloader")
root.geometry("480x360") 
#mainframe = ttk.Frame(root, padding = "300 240 300 240")
#mainframe.grid(column=0, row=0, sticky=(N, W, E, S))
root.columnconfigure(0, weight=0)
root.rowconfigure(0, weight=0)
###########################################
menubar = Menu(root)
root.config(menu=menubar)

# create the file_menu
file_menu = Menu(
    menubar,
    tearoff=0
)

# add menu items to the File menu
file_menu.add_command(label='Thêm danh sách trạm')
file_menu.add_command(label='Open...')
file_menu.add_separator()

# add Exit menu item
file_menu.add_command(
    label='Exit',
    command=root.destroy
)

# add the File menu to the menubar
menubar.add_cascade(
    label="File",
    menu=file_menu
)
# create the Help menu
help_menu = Menu(
    menubar,
    tearoff=0
)

#################################
def openHelpWindow():
    help_window = Toplevel(root)
    # Toplevel widget
    help_window.title("About")
    # sets the geometry of toplevel
    help_window.geometry("250x120")
    # A Label widget to show in toplevel
    Label(help_window,
          text ="\n\nĐược tạo bởi Đào Anh Công\n - Đài KTTV Bắc Trung Bộ")\
          .pack()

help_menu.add_command(label='About...',command = openHelpWindow)
#################################

# add the Help menu to the menubar
menubar.add_cascade(
    label="Help",
    menu=help_menu
)
###########################################
ttk.Label(root, text="Chọn loại số liệu:").\
                 place(x=10,y =20)
obs_methods = ['Quan trắc thủ công','Mưa tự động',\
           'Số liệu hồ chứa']
method = StringVar()
method.set('Quan trắc thủ công')
drop_method = OptionMenu(root,method,*obs_methods)
drop_method.place(x=110,y =15)

import datetime as dt
now = str(dt.datetime.now().date())
now1 = str(dt.datetime.now().time())[0:2]
#print (now)
year1 = int(now[0:4])
month1 = int(now[5:7])
day1 = int(now[8:10])
#print (year1, month1, day1)
####
hours = []
for i in range(0,24):
   if (i<10):
      hh = '0'+str(i)
   else:
      hh = str(i)
   hours.append(hh)
start_hour = StringVar()
start_hour.set('00')
drop_start_hour = OptionMenu(root,start_hour,*hours)
drop_start_hour.place(x=60,y =95)
#begin_hour = start_hour.get()

end_hour = StringVar()
end_hour.set(now1)
drop_end_hour = OptionMenu(root,end_hour,*hours)
drop_end_hour.place(x=270,y =95)

var_s = []
#finish_hour = end_hour.get()
################### Chọn biến
class Checkbar(Frame):
   def __init__(self, parent=None, picks=[], side=LEFT, anchor=W):
      Frame.__init__(self, parent)
      self.vars = []
      for pick in picks:
         var = IntVar()
         chk = Checkbutton(self, text=pick, variable=var)
         chk.pack(side=side, anchor=anchor, expand=NO)
         self.vars.append(var)
   def state(self):
      return map((lambda var: var.get()), self.vars)
def var_switcher(var):
    switcher = {
        'Nhiệt độ': 'AirTemp',
        'Lượng mưa': 'Precip',
        'Độ ẩm': 'RH',
        'Hướng gió':'WDir',
        'Tốc độ gió' : 'WSpeed',
        'Khí áp': 'P_SL',
        'Mực nước': 'S',
        'Lưu Lượng': 'Q',
        'Mưa tự động 6h': 'Precip/6h.total',
        'Mưa tự động 1h': 'Precip/h.total',
        'Mực nước hồ': 'ResLVL'
        }
    return switcher.get(var,"Không đúng biến")    

method_choice = ''
def select_var_window():
    variables_manual = ['Nhiệt độ','Lượng mưa','Độ ẩm','Hướng gió',\
             'Tốc độ gió','Khí áp','Mực nước','Lưu Lượng']
    variables_auto = ['Mưa tự động 6h','Mưa tự động 1h']
    variables_reservoir = ['Mực nước hồ','Lưu Lượng']
    select_var_window = Toplevel(root)
    select_var_window.title("Chọn biến")
    select_var_window.geometry("620x150")
    global method_choice
    method_choice = method.get()
    print (method_choice)
    var_choice = []
    
    if (method_choice == 'Quan trắc thủ công'):
        var_choice = variables_manual
    elif (method_choice == 'Mưa tự động'):
        var_choice = variables_auto
    elif (method_choice == 'Số liệu hồ chứa'):
        var_choice = variables_reservoir
        
    var_list = Checkbar(select_var_window, var_choice)
    var_list.pack(side=TOP,  fill=X)
    var_list.config(relief=GROOVE, bd=2)
    global begin_hour
    begin_hour = start_hour.get()
    global finish_hour
    finish_hour = end_hour.get()
    global start_date
    start_date = start_day_picker.get_date()
    global end_date
    end_date = end_day_picker.get_date()
    def var_states(): 
        state = list(var_list.state())
        global vars_name
        vars_name = []
        for i in range(len(state)):
            if (state[i] > 0):
                vars_name.append(var_choice[i])
        print (vars_name)
        global var_s
        var_s = []
        for name in vars_name:
            var_s.append(var_switcher(name))
        print (var_s)
    #for name in vars_name:
    #    var_s.append(var_switcher(name))
    Cancel_button = Button(select_var_window, text='Cancel',\
                           command=select_var_window.destroy).place(x=330, y=50)
    Ok_button = Button(select_var_window, text='Ok',\
                       command=lambda:[var_states(),select_var_window.destroy()]).\
                       place(x=250, y=50)
    
select_button = Button(root, text = "Chọn biến" , command = select_var_window)\
                .place(x=300,y=17)

from tkcalendar import DateEntry
ttk.Label(root, text="Bắt đầu:").\
                 place(x=10,y =100)
start_day_picker = DateEntry(root,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', \
                borderwidth=2)
start_day_picker.place(x=120,y =100)

####
ttk.Label(root, text="Kết thúc:").\
                 place(x=220,y =100)
end_day_picker = DateEntry(root,selectmode='day',\
                             width=8, year=year1, month=month1,\
                day=day1,background='darkblue', foreground='white', \
                borderwidth=2)
end_day_picker.place(x=330,y =100)

    
#def get_input():
#    label = Label( root , text = " ")
#    label.place(x=50,y =230)
#    variable_choice = var_switcher(Variable.get())
#    print (variable_choice)
#    region_choice = region.get()
#    begin_hour = start_hour.get()
#    finish_hour = end_hour.get()
#    global start_date
#    start_date = start_day_picker.get_date()
#    global end_date
#    end_date = end_day_picker.get_date()
#    bien = ''
#    for i in range(len(variable_choice)):
#        bien = bien + variable_choice[i]
#    label.config( text = "begin is: "+ begin_hour+' '+str(start_day) +\
#                  " end is: "+ finish_hour+' '+str(end_day)\
#                  +region_choice+bien )
    
#button = Button(root, text = "get_input" , command = get_input).place(x=200,y =200)

import urllib.request
import os
cwd = os.getcwd()

def get_manual_data():
    data_dir = os.path.join(cwd,"/Data/")
    output_dir = os.path.join(cwd,"/Output/")
    global begin_hour
    begin_hour = start_hour.get()
    global finish_hour
    finish_hour = end_hour.get()
    global start_date
    start_date = start_day_picker.get_date()
    global end_date
    end_date = end_day_picker.get_date()
    
    def show_error():
        tk.messagebox.showinfo('Xảy ra lỗi','Vui lòng chọn biến')
    
    if (method_choice == 'Quan trắc thủ công'):
        output_file = os.path.join(cwd+"/Output/",'Data_manual_'+str(start_date)+"_"+str(end_date)+".xlsx")
    elif (method_choice == 'Mưa tự động'):
        output_file = os.path.join(cwd+"/Output/",'Data_auto_'+str(start_date)+"_"+str(end_date)+".xlsx")
    elif (method_choice == 'Số liệu hồ chứa'):
        output_file = os.path.join(cwd+"/Output/",'Data_reservor_'+str(start_date)+"_"+str(end_date)+".xlsx")
    else:
        show_error()
    template_file = os.path.join(cwd+"/Output/","Template.xlsx")
    from shutil import copy2
    copy2(template_file,output_file)
    print ('Đang lấy dữ liệu')
    print (var_s)
    for name in vars_name:
        var = var_switcher(name)
        if (method_choice == 'Quan trắc thủ công'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
             +"=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/"+\
         var+"/Manual.O"
            print (url)
            stations_file =os.path.join(cwd+'/Data/',name+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request"\
               +"=getTimeseriesValues&datasource=0&format=xlsx&ts_path=5/*/"\
           +var+"/Manual.O&returnfields=Timestamp,Value&metadata=true&from="+\
           str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",name+"Manual_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        ###################
        elif (method_choice == 'Mưa tự động'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                 "&request=getTimeseriesList&datasource=0&format=xlsx&ts_path=5/*/Precip/6h.total"
            print (url)
            stations_file =os.path.join(cwd+'/Data/',name+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices"+\
                   "&request=getTimeseriesValues&datasource=0&format=xlsx&ts_path=5/*/"+var+\
                   "&returnfields=Timestamp,Value&metadata=true&"+\
                   "from="+str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",name+"Auto_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
        elif (method_choice == 'Số liệu hồ chứa'):
            url= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                 "getTimeseriesList&datasource=0&format=xlsx&site_no=11*&station_name="+\
                 "Ngan%20Truoi,Huong%20Son,Ho%20Ho,Nam%20Mo,Nam%20Non,Ban%20Ve,Ban%20Ang,"+\
                 "Khe%20Bo,Nhan%20Hac%20B,Chau%20Thang,Ban%20Mong,Ho%20Song%20Muc,Hua%20Na,"+\
                 "Dong%20Van,Cua%20Dat,Xuan%20Minh,Bai%20Thuong,Cam%20Thuy%201,Ba%20Thuoc%201"+\
                 ",Ba%20Thuoc%202,Hoi%20Xuan,Thanh%20Son,Trung%20Son&ts_name=00%20-%20Original%20"+\
                 "&parametertype_name="+var
            print (url)
            stations_file =os.path.join(cwd+'/Data/',name+"_station_"+str(start_date)+"_"+str(end_date)+".xlsx")
            print (stations_file)
            urllib.request.urlretrieve(url,stations_file)
            print ('check0')
            if (var == 'ResLVL'):
                url_2= "https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                   "getTimeseriesValues&datasource=0&format=xlsx&ts_id=308218010,308678010,310078010,"+\
                   "309878010,310968010,311048010,311018010,309918010,311028010,310048010,311718010,"+\
                   "311698010,96483010,96248010,309908010,96533010,308088010,310058010,96313010,"+\
                   "310038010,309398010,310068010,96228010&metadata=true&"+\
                   "from="+str(start_date)+"%20"+begin_hour+":00:00&to="+str(end_date)+"%20"+finish_hour+":00:00"
            else:
                url_2="https://cdh.vnmha.gov.vn/KiWIS/KiWIS?service=kisters&type=queryServices&request="+\
                       "getTimeseriesValues&datasource=0&format=xlsx&ts_id=308212010,308216010,308210010,"+\
                       "308215010,308672010,308675010,308670010,308676010,309872010,309875010,309876010,"+\
                       "310072010,310070010,310075010,309870010,310076010,310966010,310960010,310965010,"+\
                       "310962010,311010010,311015010,311012010,311040010,311046010,311042010,311016010,"+\
                       "311045010,309912010,309915010,309916010,309910010,311020010,311022010,311025010,"+\
                       "311026010,310045010,310046010,310042010,310040010,311696010,311690010,311695010,"+\
                       "311710010,311715010,311716010,311692010,311712010,96480010,96481010,96245010,"+\
                       "96246010,96251010,96478010,96243010,96486010,309900010,309905010,309902010,309906010,"+\
                       "96531010,96528010,96530010,96536010,308086010,308080010,308082010,308085010,"+\
                       "310055010,310056010,310050010,310052010,96316010,96308010,96311010,96310010,"+\
                       "310030010,310032010,310035010,310036010,309392010,309390010,309395010,309396010,"+\
                       "310062010,310066010,310065010,310060010,96226010,96231010,96223010,96225010"+\
                       "&metadata=true&"+"from="+str(start_date)+"%20"+begin_hour+":00:00&to="+\
                       str(end_date)+"%20"+finish_hour+":00:00"
            print (url_2)
            data_file = os.path.join(cwd+ "/Data/",name+"Res_data_"+str(start_date)+"_"+str(end_date)+".xlsx")
            urllib.request.urlretrieve(url_2,data_file)
            
        
        
        import openpyxl
        from openpyxl import Workbook
        from openpyxl import load_workbook
        sta_wb = load_workbook(stations_file, data_only=True)
        ## this is the part which read the stations name and id##
        sheet_sta = sta_wb['KiQS Result']
        last_cell = len(sheet_sta['A'])
        #print (last_cell)
        stations_name = []
        stations_id = []
        for i in range(2,last_cell+1):
            cell0 = 'A'+str(i)
            stations_name.append(sheet_sta[cell0].value)
            cell1 = 'B'+str(i)
            stations_id.append(sheet_sta[cell1].value)
        #stations_name = list(dict.fromkeys(stations_name))
        #stations_id = list(dict.fromkeys(stations_id))
        #print stations_name)
        #print (stations_id)
        data_wb = load_workbook(data_file, data_only=True)
        data = []
        obs_time = []
        sheet_names = data_wb.sheetnames
        #print (sheet_names)
        for i in range(len(stations_id)):
            stations_data = []
            obs_stamp = []
            for j in range (len(sheet_names)):
                if (stations_id[i] in sheet_names[j]):
                    sheet_name_data = sheet_names[j]
        #if (len(stations_id[i])<6):
         #   sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual.O"
        #else:
        #    sheet_name_data = "5_"+str(stations_id[i])+"_"+var+"_Manual"
            sheet_data = data_wb[sheet_name_data]
            last_data = len(sheet_data['A'])
            #print (sheet_name_data)
            for j in range(last_data+1-14):
                if (sheet_data.cell(row = j+14, column= 1).value) is not None:
                    obs_stamp.append(str(sheet_data.cell(row = j+14, column= 1).value))
                if (sheet_data.cell(row = j+14, column= 2).value) is not None:
                    stations_data.append(sheet_data.cell(row = j+14, column= 2).value)
            data.append(stations_data)
            obs_time.append(obs_stamp)
        #print (data)
        #print (obs_time)
        ## This is the part which write data to excel file
        data_wb = load_workbook(output_file)
        data_wb.create_sheet(name)
        sheet_0 = data_wb[name]
        for i in range (len(stations_name)):
            sheet_0.cell(row=i+2,column = 1,value = stations_name[i])
            sheet_0.cell(row=i+2,column = 2,value = stations_id[i])
        length = []
        for i in range (len(obs_time)):
            length.append(len(obs_time[i]))
        #print(max(length))
        i_max_len = length.index(max(length))
        #print (i_max_len)
    
        #print (obs_time[i_max_len])
        #print (stations_name[i_max_len])
        for j in range (len(obs_time[i_max_len])):
            time_stamp = obs_time[i_max_len][j][5:13]
            sheet_0.cell(row=1,column = j+3,value = time_stamp)
        
        for i in range (len(data)):
            for j in range(len(data[i])):
                for k in range(len(obs_time[i_max_len])):
                    if (obs_time[i][j] == obs_time[i_max_len][k]):
                        sheet_0.cell(row=i+2,column = k+3,value = data[i][j])
        data_wb.save(output_file)
    print ('Đã lấy xong dữ liệu')
    def open_file_after_done():
        MsgBox = tk.messagebox.askquestion ('Hoàn thành','Đã tải xong dữ liệu')
        if MsgBox == 'yes':
        #    root.quit()
        #else:
            os.startfile(output_file)
        
    open_file_after_done()

button = Button(root, text = "get_data" , command = get_manual_data)\
         .place(x=200,y =200)
        

root.mainloop()
