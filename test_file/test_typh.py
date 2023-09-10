## import essential libraries 
import os
from tkinter import Tk
from tkinter.filedialog import askopenfilename
## input typhoon name and make file
ty_name = input("Enter typhoon name: ")
cwd = os.getcwd()
output_file = os.path.join(cwd,ty_name+".csv")
print(f'file name is {output_file}')
if os.path.isfile(output_file):
   print("File already exist, proceed !")
else:
   with open(output_file,'w') as wt:
      wt.write("{},{},{},{},{},{},{},{},{},{},{},{},{},{}\n"\
                .format("typhoon_name","stations","obs_year","obs_month",\
                        "obs_day","obs_hour","obs_min","wind_dir","wind_speed",\
                        "slp","delta_p","gust_dir","gust_speed","sea_level"))
   wt.close()
## making array of data
stations = []
cloud_cover = []
t2m = []
tdew = []
wind_speed = []
wind_dir = []
station_pres = []
slp =[]
delta_p = []
gust_speed = []
gust_dir =[]
pmin = []
time_pmin = []
dir_max = []
wind_max = []
time_wind_max = []
sea_lev = []
sea_status = []
## convert cloud cover
def cloud_convert(str0):
   switcher = {
      '0': '0/10', '1': '0/10-1/10', '2': '2/10-3/10', '3': '4/10',
      '4': '5/10', '5': '6/10', '6': '7/10-8/10', '7': '9/10-10/10',
      '8': '10/10', '9': '-', '/': '-'
      }
   return switcher.get(str0,'invalid cloud cover')

### wind direction conversion
def dir_convert(str0):
   switcher = {
      '00': '00','02': 'NNE','05': 'NE','07': 'ENE', '09': 'E',
      '11': 'ESE', '14': 'SE', '16': 'SSE', '18': 'S',
      '20': 'SSW', '23': 'SW', '25': 'WSW', '27': 'W',
      '29': 'WNW', '32': 'NW', '34': 'NNW', '36': 'N','': ''
      }
   return switcher.get(str0,"invalid wind direction")

## Function to get position of a sub string inside a big string
def pos(str1,str2):
   pos = -1
   if str1 in str2:
      for i in range(len(str2.split())):
         if str1 == str2.split()[i]:
            pos = i
   return pos

## Get last typh
def get_last_typh(str0):
   #stations.append('48'+str.split()[2][0:3])
   pmin_exist = 0
   wind_max_exist = 0
   typh_len = len(str0.split())
   for i in range(typh_len):
      typh_len = len(str0.split())
      if (str0.split()[i][0:1] == 'A'):
         pmin_exist += 1
         pmin.append(float(str0.split()[i][1:5])/10)
      if (str0.split()[i][0:1] == 'B'):
         time_pmin.append(str0.split()[i][1:3]+':'+str0.split()[i][3:5])
      if (str0.split()[i][0:1] == 'C'):
         wind_max_exist += 1
         dir_max.append(dir_convert(str0.split()[i][1:3]))
         wind_max.append(str0.split()[i][3:5])
      if (str0.split()[i][0:1] == 'D'):
         time_wind_max.append(str0.split()[i][1:3]+':'+str0.split()[i][3:5])
   if (pmin_exist == 0):
      pmin.append('')
      time_pmin.append('')
   if (wind_max_exist == 0):
      dir_max.append('')
      wind_max.append('')
      time_wind_max.append('')
   return

## function to get data from each typh line from typh synop
def get_record_synop(str0):
   "This function get typh data"
   #print (str.split())
   typh_len = len(str0.split())
   #print(typh_len)
   stations.append(str0.split()[0])
   cloud_cover.append(cloud_convert(str0.split()[2][0:1]))
   wind_dir.append(dir_convert(str0.split()[2][1:3]))
   wind_speed.append(str0.split()[2][3:5])
   t2m.append(float(str0.split()[3][2:5])/10)
   tdew.append(float(str0.split()[4][2:5])/10)
   
   ## get station pressure
   station_pres_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == '3') and (4 < pos(str0.split()[i],str0) < 7):
         print (str0.split()[i])
         station_pres_exist += 1
         if (float(str0.split()[i][1:5]) < 1000):
            station_pres.append(1000 + float(str0.split()[i][1:5])/10)
         elif(float(str0.split()[i][1:5]) > 1000):
            station_pres.append(float(str0.split()[i][1:5])/10)
   if (station_pres_exist == 0):
      station_pres.append('')
   
   ### get sea level pressure
   pres_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == '4') and (pos(str0.split()[i],str0) > 4):
         pres_exist += 1
         if (float(str0.split()[i][1:5]) < 1000):
            slp.append(1000 + float(str0.split()[i][1:5])/10)
         elif(float(str0.split()[i][1:5]) > 1000):
            slp.append(float(str0.split()[i][1:5])/10)
   if (pres_exist == 0):
      slp.append('')
      
   ### get delta_p
   deltap_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == '5') and\
         (pos(str0.split()[i],str0) == pos('333',str0) + 1):
         deltap_exist += 1
         #print(str0.split()[i])
         if (int(str0.split()[i][0:2]) == 58):
            delta_p.append(float(str0.split()[i][2:5])/10)
         elif (int(str0.split()[i][0:2]) == 59):
            delta_p.append(0 - float(str0.split()[i][2:5])/10)
   if (deltap_exist == 0):
      delta_p.append('')

   ### get gust_wind
   ## in case of typh mistyping:
   ## case 1: 555 9****
   gust_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == '9') and\
         (pos(str0.split()[i],str0) == pos('555',str0) + 1):
         gust_exist += 1
         #print(str.split()[i])
         gust_dir.append(dir_convert(str0.split()[i][1:3]))
         gust_speed.append(str0.split()[i][3:5])
   ## case 2: 911** 915**
   for i in range(typh_len):
      if (str0.split()[i][0:3] == '915') and\
         (pos(str0.split()[i],str0) > pos('333',str0)):
         gust_exist += 1
         gust_dir.append(dir_convert(str0.split()[i][3:5]))
      if (str0.split()[i][0:3] == '911') and\
         (pos(str0.split()[i],str0) > pos('333',str0)):
         gust_speed.append(str0.split()[i][3:5])
   #print ('gust_exist = ',gust_exist)
   if (gust_exist == 0):
      gust_dir.append('')
      gust_speed.append('')

   ### get water level
   sea_lev_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == 'E'):
         sea_lev_exist += 1
         sea_lev.append(str0.split()[i][1:4])
         sea_status.append(str0.split()[i][4:5])
   if (sea_lev_exist == 0):
      sea_lev.append('')
      sea_status.append('')
   get_last_typh(str0)
   return
# this function get data from typh thuy van
def get_record_TV(str0):
   typh_len = len(str0.split())
   stations.append(str0.split()[2])
   cloud_cover.append(cloud_convert(str0.split()[3][0:1]))
   wind_dir.append(dir_convert(str0.split()[3][1:3]))
   wind_speed.append(str0.split()[3][3:5])
   t2m.append('')
   tdew.append('')
   station_pres.append('')
   slp.append("")
   delta_p.append("")
   sea_lev.append('')
   if (typh_len > 5):
      gust_dir.append(dir_convert(str0.split()[4][1:3]))
      gust_speed.append(str0.split()[4][3:5])
   else:
      gust_dir.append("")
      gust_speed.append("")
   sea_lev_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == 'E'):
         sea_lev_exist += 1
         sea_lev.append(str0.split()[i][1:4])
         sea_status.append(str0.split()[i][4:5])
   if (sea_lev_exist == 0):
      sea_lev.append('')
      sea_status.append('')
   get_last_typh(str0)
   return

## function to get data from each typh line from typh KT
def get_record_typh(str0):
   "This function get typh data"
   #print (str.split())
   typh_len = len(str0.split())
   #print(typh_len)
   stations.append('48'+str0.split()[2][0:3])
   cloud_cover.append(cloud_convert(str0.split()[3][0:1]))
   wind_dir.append(dir_convert(str0.split()[3][1:3]))
   wind_speed.append(str0.split()[3][3:5])
   t2m.append(float(str0.split()[4][2:5])/10)
   tdew.append(float(str0.split()[5][2:5])/10)
   ## get station_pressure
   #print(str0.split()[2][0:3])
   if (typh_len > 10):
      if (float(str0.split()[6][1:5]) < 1000):
         station_pres.append(1000+float(str0.split()[6][1:5])/10)
      elif(float(str0.split()[6][1:5]) > 1000):
         station_pres.append(float(str0.split()[6][1:5])/10)
   else:
      station_pres.append("")
   ## get slp
   if (typh_len > 10):
      if (float(str0.split()[7][1:5]) < 1000):
         slp.append(1000+float(str0.split()[7][1:5])/10)
      elif(float(str0.split()[7][1:5]) > 1000):
         slp.append(float(str0.split()[7][1:5])/10)
   else:
      slp.append("")
   ## get delta P
   if (typh_len > 10):
      if (int(str0.split()[8][0:2]) == 58):
           delta_p.append(float(str0.split()[8][2:5])/10)
      elif (int(str0.split()[8][0:2]) == 59):
           delta_p.append(0 - float(str0.split()[8][2:5])/10)
   else:
      delta_p.append("")
   ## get gust wind
   gust_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == '9'):
         gust_exist += 1
         print(str0.split()[2][0:3])
         print(str0.split()[i])
         gust_dir.append(dir_convert(str0.split()[i][1:3]))
         gust_speed.append(str0.split()[i][3:5])
   if (gust_exist == 0):
      gust_dir.append('')
      gust_speed.append('')
   ## get water level
   sea_lev_exist = 0
   for i in range(typh_len):
      if (str0.split()[i][0:1] == 'E'):
         sea_lev_exist += 1
         sea_lev.append(str0.split()[i][1:4])
         sea_status.append(str0.split()[i][4:5])
   if (sea_lev_exist == 0):
      sea_lev.append('')
      sea_status.append('')
   get_last_typh(str0)
   return

## Get data from input file
Tk().withdraw() # disable the root windows
input_file = askopenfilename() # open file browsing dialog box
#print(input_file)

## Read raw data from file
import datetime
with open(input_file,'r') as dt:
   lines = dt.read()
   record_temp = lines.split("\n")
   #print(record_temp)
   for i in range(record_temp.count('')):
      record_temp.remove('')
   typh_type = (record_temp[2].split()[0]).lower() #find obs typh type aaxx or typh
   print (typh_type)
   obs_time = record_temp[1].split()[2]     #get the obs time
   print (obs_time)
   obs_year = datetime.datetime.now().year   #define year and month
   obs_month = datetime.datetime.now().month
   obs_day =  obs_time[0:2]
   obs_hour = obs_time[2:4]
   obs_min = obs_time[4:6]
   dt.close()
# print(obs_year, obs_month, obs_day, obs_hour, obs_min)
record_typh = []

for i in range(len(record_temp)):
   if (len(record_temp[i].split()) >= 5):
      record_typh.append(record_temp[i])

## get data from raw data
################
for i in range(len(record_typh)):
   if (typh_type == 'typh0' or typh_type == 'typh1'):
      get_record_typh(record_typh[i])
   elif (typh_type == 'typh2' or typh_type == 'typh3'):
      get_record_TV(record_typh[i])
   else:
      get_record_synop(record_typh[i])
################
      
## Write data to file
def overwrite_record(file):
   dt = open(output_file,'r')
   record_temp = dt.read()
   lines = record_temp.split("\n")
   dt.close()
   #print (lines)
   for i in range(len(stations)):
      check_line = f'{ty_name},{stations[i]},{obs_year},{obs_month},\
{obs_day},{obs_hour},{obs_min}'
      write_line = "{},{},{},{},{},{},{},{},{},{},{},{},{},{}"\
            .format(ty_name,stations[i],obs_year,obs_month,obs_day,\
                     obs_hour,obs_min,wind_dir[i],wind_speed[i],slp[i],\
                     delta_p[i],gust_dir[i],gust_speed[i],sea_lev[i])
      for k in range(0,len(lines)):
         if check_line in lines[k]:
            #print ('this line is replacable')
            lines[k] = write_line
            break
   
   with open('typh_temp.csv','w') as wt:
      for i in range(len(lines)):
         #print(lines[i])
         wt.write(lines[i] + '\n')
   wt.close()
   import os
   from shutil import copy2
   os.remove(output_file)
   copy2('typh_temp.csv',output_file)
   return

def write_record(file):
   with open(file,'a+') as wt:
      for i in range(len(stations)):
         wt.write("{},{},{},{},{},{},{},{},{},{},{},{},{},{}\n"\
            .format(ty_name,stations[i],obs_year,obs_month,\
                    obs_day,obs_hour,obs_min,wind_dir[i],\
                    wind_speed[i],slp[i],delta_p[i],gust_dir[i],\
                    gust_speed[i],sea_lev[i]))
   wt.close()
   return

# Write or overwrite data file
check_line = f'{ty_name},{stations[i]},{obs_year},{obs_month},\
{obs_day},{obs_hour},{obs_min}'
print (check_line)
rt = open(output_file,'r')
line_temp = rt.read()
#print (line_temp)
rt.close()

print (gust_dir)
if check_line in line_temp:
   print ('This file is replacable')
   print ('******************')
   overwrite_record(output_file)
else:
   write_record(output_file)
   
##################################################################

# Function to Write or update data into database
import pyodbc
import os
mdb_file = os.path.join(cwd,"DBdienbaoKTTVHV.accdb")
##[x for x in pyodbc.drivers() if x.startswith('Microsoft Access Driver')]
print(mdb_file)
mdb_driver = '{Microsoft Access Driver (*.mdb, *.accdb)}'
conn = pyodbc.connect('DRIVER={}; DBQ={}'.format(mdb_driver,mdb_file))
cursor = conn.cursor()
### function to write data to database
def write_data_to_db():
   for i in range(len(stations)):
      command_test= "select 1 from TYPH where Typh_name='{}' and \
STNO='{}' and yYear='{}' and mMonth='{}' and dDay='{}' and hHour='{}' \
and mMin='{}'".format(ty_name,stations[i],obs_year,obs_month,obs_day,obs_hour,\
                  obs_min)
      #print (command_test)
      command0= "insert into TYPH (Typh_name,STNO,yYear,mMonth,dDay,hHour,\
mMin,Typh_type,cloud_cover,wind_dir,wind_speed,TTT,TdTdTd,station_pres,slp,\
deltaP24,gust_dir,gust_speed,ApnPn,BGGPP,CDxDx,CFxFx,DGGPP,Ehhh,Ez)\n\
values ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',\
'{}','{}','{}','{}','{}','{}','{}','{}','{}','{}',\
'{}','{}')".format(ty_name,stations[i],obs_year,obs_month,obs_day,obs_hour,\
               obs_min,typh_type,cloud_cover[i],wind_dir[i],wind_speed[i],\
               t2m[i],tdew[i],station_pres[i],slp[i],delta_p[i],gust_dir[i],\
               gust_speed[i],pmin[i],time_pmin[i],dir_max[i],wind_max[i],\
               time_wind_max[i],sea_lev[i],sea_status[i])
      #print(command0)
      command1="update TYPH\n\
set Typh_type='{}',cloud_cover='{}',wind_dir='{}',wind_speed='{}',\
TTT='{}',TdTdTd='{}',station_pres='{}',slp='{}',deltaP24='{}',gust_dir='{}',\
gust_speed='{}',ApnPn='{}',BGGPP='{}',CDxDx='{}',CFxFx='{}',DGGPP='{}',\
Ehhh='{}',Ez='{}'\n\
where Typh_name='{}' and STNO='{}' and yYear='{}' and mMonth='{}' and \
dDay='{}' and hHour='{}' and \
mMin='{}'".format(typh_type,cloud_cover[i],wind_dir[i],wind_speed[i],\
               t2m[i],tdew[i],station_pres[i],slp[i],delta_p[i],gust_dir[i],\
               gust_speed[i],pmin[i],time_pmin[i],dir_max[i],wind_max[i],\
               time_wind_max[i],sea_lev[i],sea_status[i],ty_name,stations[i],\
               obs_year,obs_month,obs_day,obs_hour,obs_min)
      #print(command1)
      #cursor.execute(command0)
      cursor.execute(command_test)
      test = cursor.fetchone()
      #print (type(test))
      #print (test)
      if not test:
         print ('record not exist, make new record')
         cursor.execute(command0)
      else:
         print ('record exist, update record')
         cursor.execute(command1)
      conn.commit()
      #cursor.close()
   return

# Function to write data to excel file
typh_stations=['48842','48/67','48/68','48/69','48840','48/70','48/72',\
          '48/66','48/74','48844','48/75','48/76','48/79','48/77',\
          '48/80','48/81','48845','48/82','48846','48/84','48/73','48/86',\
          '73417','72436','72446']
def get_data_db(ty_day,ty_hour,ty_minute):
   ################
   for station in typh_stations:
      #print (ty_day,ty_hour,ty_minute)
      command2= "select hHour,mMin,wind_dir,wind_speed,slp,deltaP24,gust_dir,\
gust_speed,ApnPn,BGGPP,CDxDx,CFxFx,DGGPP from TYPH where Typh_name='{}' and \
STNO='{}' and yYear='{}' and mMonth='{}' and dDay='{}' and hHour='{}' and \
mMin='{}'".format(ty_name,station,obs_year,obs_month,ty_day,ty_hour,ty_minute)
      cursor.execute(command2)
      record = cursor.fetchall()
      #print(list(record))
      if not (list(record)):
         typh_hour0.append('')
         typh_min0.append('')
         wind_dir0.append('')
         wind_speed0.append('')
         slp0.append('')
         delta_p0.append('')
         gust_dir0.append('')
         gust_speed0.append('')
         pmin0.append('')
         time_pmin0.append('')
         dir_max0.append('')
         wind_max0.append('')
         time_wind_max0.append('')
      else:
         for row in record:
            if not (list(row)[0]):
               typh_hour0.append('')
            else:
               typh_hour0.append(list(row)[0])
            if not (list(row)[1]):
               typh_min0.append('')
            else:
               typh_min0.append(list(row)[1])
            if not (list(row)[2]):
               wind_dir0.append('')
            else:
               wind_dir0.append(list(row)[2])
            if not (list(row)[3]):
               wind_speed0.append('')
            else:
               wind_speed0.append(int(list(row)[3]))
            if not (list(row)[4]):
               slp0.append('')
            else:
               slp0.append(float(list(row)[4]))
            if not (list(row)[5]):
               delta_p0.append('')
            else:
               delta_p0.append(float(list(row)[5]))
            if not (list(row)[6]):
               gust_dir0.append('')
            else:
               gust_dir0.append(list(row)[6])
            if not (list(row)[7]):
               gust_speed0.append('')
            else:
               gust_speed0.append(int(list(row)[7]))
            if not (list(row)[8]):
               pmin0.append('')
            else:
               pmin0.append(float(list(row)[8]))
            if not (list(row)[9]):
               time_pmin0.append('')
            else:
               time_pmin0.append(list(row)[9])
            if not (list(row)[10]):
               dir_max0.append('')
            else:
               dir_max0.append(list(row)[10])
            if not (list(row)[11]):
               wind_max0.append('')
            else:
               wind_max0.append(list(row)[11])
            if not (list(row)[12]):
               time_wind_max0.append('')
            else:
               time_wind_max0.append(list(row)[12])
   return
###################
## Write data to db and excel file
write_data_to_db()
###################

#print(slp)
### Write data to excel file
output_excel_file = os.path.join(cwd,'TYPH_' + ty_name + '.xlsx')
import os.path
from shutil import copy2
if os.path.isfile(output_excel_file):
   print('Excel file already exist')
else:
   print('Excel File doesnt exist, copy new file')
   copy2('TYPH_BAO_CHUAN.xlsx',output_excel_file)
###################################

## get all days that have typh
get_typh_day = "select dDay from Typh where Typh_name='{}' and \
yYear='{}' and mMonth='{}'".format(ty_name,obs_year,obs_month,)
cursor.execute(get_typh_day)
typh_swap0 = cursor.fetchall()
typh_swap1 = [x[0] for x in typh_swap0]
typh_time_days = sorted(list(set(typh_swap1)))
print (typh_time_days)

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
wb = load_workbook(output_excel_file)
sheet0 = wb['DATA']
sheet0['I1'] = 'TYPH ' + ty_name.upper()
i = 0
## clear old data before writing in excel file
for row in sheet0['H7:BN106']:
   for cell in row:
      cell.value = None
## start writing data to excel file
for ty_day in typh_time_days:
   ## get all hour time that have typh
   get_typh_hour = "select hHour from Typh where Typh_name='{}' and \
yYear='{}' and mMonth='{}' and dDay='{}'".format(ty_name,obs_year,\
                                                 obs_month,ty_day)
   cursor.execute(get_typh_hour)
   typh_swap0 = cursor.fetchall()
   typh_swap1 = [x[0] for x in typh_swap0]
   typh_time_hour = sorted(list(set(typh_swap1)))
   for ty_hour in typh_time_hour:  # Write hour
      get_typh_min = "select mMin from Typh where Typh_name='{}' and \
yYear='{}' and mMonth='{}' and dDay='{}' and hHour='{}'".format(ty_name,\
                                             obs_year,obs_month,\
                                             ty_day,ty_hour)
      cursor.execute(get_typh_min)
      typh_swap_0 = cursor.fetchall()
      typh_swap_1 = [x[0] for x in typh_swap_0]
      typh_time_min = sorted(list(set(typh_swap_1)))
      for ty_minute in (typh_time_min):
      ############################
         print (ty_day+' '+ty_hour +':'+ ty_minute)
         ##### important_part here
         ## making array of data
         typh_hour0 = []
         typh_min0 = []
         wind_speed0 = []
         wind_dir0 = []
         station_pres0 = []
         slp0 =[]
         delta_p0 = []
         gust_speed0 = []
         gust_dir0 =[]
         pmin0 = []
         time_pmin0 = []
         dir_max0 = []
         wind_max0 = []
         time_wind_max0 = []
         ###########################
         get_data_db(ty_day,ty_hour,ty_minute)
         # 8 is the starting column
         # Print days and hour of the typh
         
         sheet0.cell(row= 5, column= 8 + i, value= ty_day+'/'+str(obs_month))
         sheet0.cell(row= 6, column= 8 + i, value= ty_hour +':'+ ty_minute)

         ## Begin to write typh data into 
         for k in range(len(slp0)):
            #print (slp0[k])
            ## write slp
            if slp0[k] != "":
               sheet0.cell(row=(7 + 4*k), column= 8 + i, value=slp0[k])
            ## write delta_p
            if delta_p0[k] != "":
               sheet0.cell(row=(8 + 4*k), column= 8 + i, value=delta_p0[k])
            ## Write wind dir & wind speed
            if wind_dir0[k] != "":
               sheet0.cell(row=(9 + 4*k), column= 8 + i, value=wind_dir0[k]+' '+\
                        str(wind_speed0[k]))
            ## Write gust dir and gust speed
            if gust_dir0[k] != "":
               sheet0.cell(row=(10 + 4*k), column= 8 + i, value=gust_dir0[k]+' '+\
                     str(gust_speed0[k]))
            ## Write last typh
            if pmin0[k] != '':
               sheet0.cell(row=(7 + 4*k), column= 2, value= pmin0[k])
            ## Write time pmin
            if time_pmin0[k] != '':
               sheet0.cell(row=(8 + 4*k), column= 2, value=time_pmin0[k])
            ## write wind_max
            if dir_max0[k] != '':
               sheet0.cell(row=(9 + 4*k), column= 2, value=dir_max0[k]+' '+\
                        str(wind_max0[k]))
            ## write time wind max
            if time_wind_max0[k] != '':
               sheet0.cell(row=(10 + 4*k), column= 2, value=time_wind_max0[k])
         i += 1

wb.save(output_excel_file)

